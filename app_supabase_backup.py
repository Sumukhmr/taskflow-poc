from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from supabase import create_client, Client
from dotenv import load_dotenv
from datetime import datetime
import uuid
import threading
import time
import os
import io
import csv
import json
import requests as http_requests

load_dotenv()

app = Flask(__name__)
CORS(app)

# ── CONFIGURATION ───────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

# ── SUPABASE CLIENT ─────────────────────────────────────────────
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── APP VERSION (bump this after every deploy) ───────────────────
APP_VERSION = '1.0.0'

# ── CONSTANTS ───────────────────────────────────────────────────
# Built-in defaults — seed values only. FUNCTIONS is rebuilt at runtime
# from Supabase (teams + sub_categories tables) so the user can add their
# own functions and sub-categories without code changes.
DEFAULT_FUNCTIONS = {
    "Sales":       ["Enterprise Pipeline", "Outbound", "Inbound", "Account Management", "Partnerships"],
    "Content":     ["Pedagogy", "Content Ingestion", "Content Review/Testing", "Curriculum Design"],
    "Engineering": ["Backend", "Frontend", "Infrastructure/DevOps", "Mobile", "Data"],
    "Operations":  ["Customer Onboarding", "Process", "Vendor Management", "Internal Tools"],
    "Marketing":   ["Campaigns", "Events/Conferences", "Content Marketing", "SEO", "Social", "Brand"],
    "Finance":     ["Accounting", "Budgeting", "Payroll", "Reporting"],
    "HR":          ["Recruiting", "People Ops", "Learning & Development", "Culture"],
    "Product":     ["Feature", "Enhancement", "UI/UX", "Bug/Issue", "Analytics/Reporting", "Research"],
    "Legal":       ["Contracts", "Compliance", "IP"],
    "Support":     ["Customer Support", "Technical Support"],
}
# FUNCTIONS is mutated in place by sync_functions() so existing references
# (e.g. FUNCTIONS["Engineering"]) keep working.
FUNCTIONS = {k: list(v) for k, v in DEFAULT_FUNCTIONS.items()}
DEFAULT_TEAMS = list(DEFAULT_FUNCTIONS.keys())
VALID_STATUSES = [
    "todo", "just", "progress", "hold", "blocked", "review", "done", "reopen",
]
VALID_PRIORITIES = ["high", "medium", "low"]
VALID_QUADRANTS = ["q1", "q2", "q3", "q4"]
# Deployment environment for a task. Empty string = unset.
VALID_ENVS = ["", "local", "dev", "uat", "cloud"]

# ── IN-MEMORY CACHE ─────────────────────────────────────────────
cache = {
    "tasks": [],
    "teams": [],
    "users": [],
    "activity": [],
    "initiatives": [],
    "functions": FUNCTIONS,
    "last_sync": None
}
cache_lock = threading.Lock()

# Initiative status + accent-color whitelists. Anything outside these
# falls back to a safe default rather than rejecting the row.
VALID_INITIATIVE_STATUSES   = ["planning", "active", "done", "archived"]
VALID_INITIATIVE_COLORS     = ["pink", "violet", "amber", "sky", "emerald"]
VALID_INITIATIVE_PRIORITIES = ["P0", "P1", "P2", "P3", "P4"]
# Per-task work classification. Empty string = unset; any task from any
# function can carry one of these — no team-based restriction.
VALID_WORK_TYPES = ["", "feature", "enhancement", "issue"]

# ── TESTER AUTO-ASSIGNMENT ───────────────────────────────────────
# When a task in a dev-team function is marked done, the designated
# tester is automatically appended to assigned_to (if not already there).
DEVELOPER_TESTER_MAP = {
    "chandan mk":        "Shashikumar",
    "sankalpa":          "Shashikumar",
    "gayathri priya c v": "Shashikumar",
}
TESTER_AUTO_ASSIGN_FUNCTIONS = ["Engineering", "Development"]

# ── HELPERS ──────────────────────────────────────────────────────
def generate_task_id():
    return f"TSK-{uuid.uuid4().hex[:6].upper()}"

def generate_initiative_id():
    return f"INIT-{uuid.uuid4().hex[:6].upper()}"

def _auto_assign_tester(task):
    """Append the designated tester to assigned_to when a dev-function task
    moves to 'done'.  Returns (new_assigned_to, tester_name) on a change,
    or (None, None) when no action is needed."""
    team = (task.get("team") or "").strip()
    if team not in TESTER_AUTO_ASSIGN_FUNCTIONS:
        return None, None
    raw   = task.get("assigned_to") or ""
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    parts_lower = [p.lower() for p in parts]
    tester = None
    for dev_key, tester_name in DEVELOPER_TESTER_MAP.items():
        if dev_key in parts_lower:
            tester = tester_name
            break
    if not tester:
        return None, None
    if tester.lower() in parts_lower:
        return None, None          # already assigned
    parts.append(tester)
    return ", ".join(parts), tester

def safe_initiative(i):
    """Normalize an initiative dict — sane defaults for legacy / partial rows."""
    status = (i.get("status") or "planning").strip().lower()
    if status not in VALID_INITIATIVE_STATUSES:
        status = "planning"
    accent = (i.get("accent_color") or "violet").strip().lower()
    if accent not in VALID_INITIATIVE_COLORS:
        accent = "violet"
    priority = (i.get("priority") or "P2").strip().upper()
    if priority not in VALID_INITIATIVE_PRIORITIES:
        priority = "P2"
    return {
        "id":           i.get("id", "") or "",
        "name":         i.get("name", "") or "",
        "type":         i.get("type", "") or "",
        "description":  i.get("description", "") or "",
        "start_date":   i.get("start_date", "") or "",
        "end_date":     i.get("end_date", "") or "",
        "status":       status,
        "priority":     priority,
        "created_by":   i.get("created_by", "") or "",
        "accent_color": accent,
        "created_at":   i.get("created_at", "") or "",
        "updated_at":   i.get("updated_at", "") or "",
    }

def get_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_date(value):
    """Convert common user-typed / spreadsheet date formats to YYYY-MM-DD.

    Handles M/D/YYYY, MM-DD-YYYY, DD/MM/YYYY, DD-MM-YYYY, YYYY/M/D,
    M/D/YY, DD/MM/YY, and the canonical YYYY-MM-DD. Returns "" for empty
    input. If parsing fails for every format, returns the original string
    so the caller can decide what to do (we don't silently drop the
    user's data)."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    # Already ISO? Trust it.
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return s
    except ValueError:
        pass

    # Try common formats in order. NOTE: M/D/YYYY and DD/MM/YYYY are
    # ambiguous when day <= 12, so we try US-style first (more common
    # in spreadsheets exported from Excel/Google Sheets in en-US locale),
    # then European-style as a fallback.
    formats = [
        "%m/%d/%Y", "%m-%d-%Y",
        "%d/%m/%Y", "%d-%m-%Y",
        "%Y/%m/%d",
        "%m/%d/%y", "%d/%m/%y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return s  # fall back to original — unparseable but preserved

def write_in_background(fn, *args, **kwargs):
    def wrapper():
        try:
            fn(*args, **kwargs)
        except Exception as e:
            print(f"Background write failed: {e}")
    thread = threading.Thread(target=wrapper, daemon=True)
    thread.start()

# ── GOOGLE SHEETS BACKUP (concurrent with Supabase) ─────────────
# The Sheets write is fire-and-forget: it never blocks the API
# response and never propagates failures. If credentials.json is
# missing, SPREADSHEET_ID is empty/placeholder, or gspread isn't
# installed, we silently no-op after the first probe.
try:
    import gspread as _gspread
    from google.oauth2.service_account import Credentials as _GSCredentials
    _GSPREAD_AVAILABLE = True
except Exception:
    _GSPREAD_AVAILABLE = False

CREDENTIALS_FILE = "credentials.json"
SPREADSHEET_ID   = "12dzMiqpQQJqCZo4bIz8QQRTXaC6ynQRs7nxnCMwTnb0"

# Column layout of the backup sheet, in order. Task_ID is column A —
# the row-lookup helper depends on that position for update / delete.
_SHEET_COLUMNS = [
    "task_id", "title", "description", "assigned_to", "team", "due_date",
    "priority", "status", "created_by", "created_at", "updated_at",
    "sub_category", "quadrant", "notes", "type", "start_date",
    "attachments", "parent_task_id", "env", "comments", "depends_on",
    "milestone_marker", "initiative_id", "work_type", "status_log",
]

_sheets_lock = threading.Lock()
_sheets_spreadsheet = None                # cached gspread Spreadsheet handle
_sheets_worksheets  = {}                  # tab_name → worksheet (or None if that tab is missing)
_sheets_probed_and_failed = False         # process-wide auth-failure flag

def _get_sheets_worksheet(tab_name):
    """Return the named worksheet on the configured spreadsheet, or None.
    Lazy-connects once (auth is cached process-wide), then caches each tab
    handle. An auth-level failure disables the entire backup. A missing-tab
    failure only disables that one tab so the others keep working."""
    global _sheets_spreadsheet, _sheets_probed_and_failed
    if not _GSPREAD_AVAILABLE:                                           return None
    if not SPREADSHEET_ID or SPREADSHEET_ID == "YOUR_SHEET_ID_HERE":       return None
    if not os.path.exists(CREDENTIALS_FILE):                              return None
    if _sheets_probed_and_failed:                                         return None
    with _sheets_lock:
        if tab_name in _sheets_worksheets:
            return _sheets_worksheets[tab_name]     # may be None (previously-missing tab)
        # First-touch auth.
        if _sheets_spreadsheet is None:
            try:
                scopes = ["https://www.googleapis.com/auth/spreadsheets"]
                creds  = _GSCredentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
                client = _gspread.authorize(creds)
                _sheets_spreadsheet = client.open_by_key(SPREADSHEET_ID)
            except Exception as e:
                print(f"[sheets] auth failed, backup disabled: {e}")
                _sheets_probed_and_failed = True
                return None
        # Then open the requested tab.
        try:
            ws = _sheets_spreadsheet.worksheet(tab_name)
            _sheets_worksheets[tab_name] = ws
            print(f"[sheets] connected to {tab_name} tab successfully")
            return ws
        except Exception as e:
            print(f"[sheets] worksheet '{tab_name}' unavailable: {e}")
            _sheets_worksheets[tab_name] = None      # cache the miss so we don't retry
            return None

def get_sheets_client():
    """Backward-compatible alias — returns the Tasks worksheet."""
    return _get_sheets_worksheet("Tasks")

def _sheets_col_letter(n):
    """1 → 'A', 21 → 'U', 27 → 'AA'."""
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out

def _sheets_task_row(t):
    """Serialize a task dict into a list of strings matching _SHEET_COLUMNS."""
    def flat(col, v):
        if col == "status_log":
            return json.dumps(v, separators=(",", ":")) if v else "[]"
        if v is None:              return ""
        if isinstance(v, bool):    return "TRUE" if v else "FALSE"
        if isinstance(v, list):    return ";".join(str(x) for x in v if x is not None)
        return str(v)
    return [flat(col, (t or {}).get(col, "")) for col in _SHEET_COLUMNS]

def _sheets_find_row(ws, task_id):
    """Return the 1-indexed row containing task_id in column A, or None."""
    if not task_id:
        return None
    try:
        col_a = ws.col_values(1)
        for i, v in enumerate(col_a):
            if v == task_id:
                return i + 1
    except Exception as e:
        print(f"[sheets] col_values failed: {e}")
    return None

def _sheets_apply(ws, action, task):
    tid = (task or {}).get("task_id", "")
    row = _sheets_task_row(task)
    end_col = _sheets_col_letter(len(_SHEET_COLUMNS))
    if action == "insert":
        ws.append_row(row, value_input_option="USER_ENTERED")
    elif action == "update":
        r = _sheets_find_row(ws, tid)
        if r:
            ws.update(range_name=f"A{r}:{end_col}{r}", values=[row], value_input_option="USER_ENTERED")
        else:
            # Row not present yet (e.g. the task predates the backup) —
            # append so subsequent updates can find it.
            ws.append_row(row, value_input_option="USER_ENTERED")
    elif action == "delete":
        r = _sheets_find_row(ws, tid)
        if r:
            ws.delete_rows(r)

def backup_task_to_sheets(action, task):
    """Fire-and-forget backup of a single task to Google Sheets.
    Runs on its own daemon thread — the API response never waits.
    Every exception is caught + printed; nothing propagates."""
    def worker():
        try:
            ws = get_sheets_client()
            if ws is None:
                return
            _sheets_apply(ws, action, task)
        except Exception as e:
            print(f"[sheets backup] {action} for {(task or {}).get('task_id','?')} failed: {e}")
    threading.Thread(target=worker, daemon=True).start()

def backup_tasks_to_sheets(action, tasks):
    """Bulk variant — one background thread processes all N tasks
    sequentially, so bulk-add doesn't spawn dozens of concurrent
    connections to Google."""
    def worker():
        try:
            ws = get_sheets_client()
            if ws is None:
                return
            for t in tasks:
                try:
                    _sheets_apply(ws, action, t)
                except Exception as e:
                    print(f"[sheets bulk {action}] {(t or {}).get('task_id','?')} failed: {e}")
        except Exception as e:
            print(f"[sheets bulk] {action} outer failure: {e}")
    threading.Thread(target=worker, daemon=True).start()


# ── Sheets backup: Users / Teams / sub_categories ─────────────
# Same pattern as tasks: daemon thread, silent failure, no-op when
# not configured. Each tab has its own column layout + primary key
# (first column) used to locate rows for update / delete.

_USER_COLUMNS   = ["name", "team", "email", "password", "role"]
_SUBCAT_COLUMNS = ["id", "function", "sub_category", "created_at"]

def _sheets_row(record, columns):
    """Generic flatten — same rules as _sheets_task_row."""
    def flat(v):
        if v is None:               return ""
        if isinstance(v, bool):     return "TRUE" if v else "FALSE"
        if isinstance(v, list):     return ";".join(str(x) for x in v if x is not None)
        return str(v)
    return [flat((record or {}).get(col, "")) for col in columns]

def _sheets_apply_generic(ws, action, key, row, columns):
    """Shared insert/update/delete logic for any tab keyed by column A."""
    end_col = _sheets_col_letter(len(columns))
    if action == "insert":
        ws.append_row(row, value_input_option="USER_ENTERED")
    elif action == "update":
        r = _sheets_find_row(ws, key)
        if r:
            ws.update(range_name=f"A{r}:{end_col}{r}", values=[row],
                      value_input_option="USER_ENTERED")
        else:
            # Row not present yet — treat update as an append so subsequent
            # updates can find it.
            ws.append_row(row, value_input_option="USER_ENTERED")
    elif action == "delete":
        r = _sheets_find_row(ws, key)
        if r:
            ws.delete_rows(r)

def backup_user_to_sheets(action, user):
    """Backs a user row up to the 'Users' tab. Row identity = Name (col A)."""
    def worker():
        try:
            ws = _get_sheets_worksheet("Users")
            if ws is None:
                return
            key = (user or {}).get("name", "")
            row = _sheets_row(user, _USER_COLUMNS)
            _sheets_apply_generic(ws, action, key, row, _USER_COLUMNS)
        except Exception as e:
            print(f"[sheets backup user] {action} for {(user or {}).get('name','?')} failed: {e}")
    threading.Thread(target=worker, daemon=True).start()

def backup_team_to_sheets(action, team):
    """Backs a team up to the 'Teams' tab. `team` may be a bare string or
    a dict like {'team_name': '…'}. Row identity = Team_Name (col A)."""
    def worker():
        try:
            ws = _get_sheets_worksheet("Teams")
            if ws is None:
                return
            if isinstance(team, str):
                name = team
            else:
                name = (team or {}).get("team_name") or (team or {}).get("name") or ""
            if not name:
                return
            _sheets_apply_generic(ws, action, name, [name], ["team_name"])
        except Exception as e:
            print(f"[sheets backup team] {action} failed: {e}")
    threading.Thread(target=worker, daemon=True).start()

def backup_subcategory_to_sheets(action, subcategory):
    """Backs a sub-category up to the 'sub_categories' tab. Row identity =
    ID (col A). ID is either the numeric primary key from Supabase or, for
    freshly-inserted rows where we don't know the ID yet, we fall back to
    a synthetic key of 'function::sub_category' — that keeps update/delete
    functional if the Supabase insert happens after the Sheets write."""
    def worker():
        try:
            ws = _get_sheets_worksheet("sub_categories")
            if ws is None:
                return
            sc = subcategory or {}
            key = str(sc.get("id") or f"{sc.get('function','')}::{sc.get('sub_category','')}")
            # Normalise the record so the row always has the composite key
            # in column A regardless of whether Supabase returned an id.
            rec = dict(sc)
            rec["id"] = key
            row = _sheets_row(rec, _SUBCAT_COLUMNS)
            _sheets_apply_generic(ws, action, key, row, _SUBCAT_COLUMNS)
        except Exception as e:
            print(f"[sheets backup subcategory] {action} failed: {e}")
    threading.Thread(target=worker, daemon=True).start()


def log_activity(user, action, task_id="", task_title="", details=""):
    row = {
        "timestamp": get_timestamp(),
        "username": user or "System",
        "action": action,
        "task_id": task_id,
        "task_title": task_title,
        "details": details
    }
    with cache_lock:
        cache.setdefault("activity", []).insert(0, row)
        cache["activity"] = cache["activity"][:200]

    def sheet_write():
        supabase.table("activity_log").insert(row).execute()
    write_in_background(sheet_write)

def safe_task(t):
    """Ensure task dict has all fields with proper defaults."""
    return {
        "task_id":      t.get("task_id", ""),
        "title":        t.get("title", ""),
        "description":  t.get("description", "") or "",
        "assigned_to":  t.get("assigned_to", "") or "",
        "team":         t.get("team", "") or "",
        "due_date":     t.get("due_date", "") or "",
        "priority":     t.get("priority", "medium") or "medium",
        "status":       t.get("status", "todo") or "todo",
        "created_by":   t.get("created_by", "") or "",
        "created_at":   t.get("created_at", "") or "",
        "updated_at":   t.get("updated_at", "") or "",
        "sub_category": t.get("sub_category", "") or "",
        "quadrant":     t.get("quadrant", "q2") or "q2",
        "notes":        t.get("notes", "") or "",
        "start_date":     t.get("start_date", "") or "",
        "attachments":    t.get("attachments") or [],
        "parent_task_id": t.get("parent_task_id", "") or "",
        "env":            (t.get("env", "") or "").strip().lower(),
        "comments":       t.get("comments") or [],
        # Timeline view: explicit task dependencies + milestone marker.
        "depends_on":       t.get("depends_on") or [],
        "milestone_marker": bool(t.get("milestone_marker") or False),
        "initiative_id":    (t.get("initiative_id") or "").strip(),
        "work_type":        (t.get("work_type") or "").strip().lower(),
        "status_log":       t.get("status_log") or [],
    }

# ── CACHE SYNC ──────────────────────────────────────────────────
def _fetch_all_pages(table_name, select="*", order_col=None, order_desc=False, page_size=1000):
    """Fetch every row from a Supabase table, bypassing the default 1000-row cap."""
    rows = []
    offset = 0
    while True:
        q = supabase.table(table_name).select(select)
        if order_col:
            q = q.order(order_col, desc=order_desc)
        res = q.range(offset, offset + page_size - 1).execute()
        batch = res.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def sync_cache_from_supabase():
    with cache_lock:
        last = cache.get("last_sync")
    if last:
        try:
            from datetime import datetime as _dt
            elapsed = (_dt.now() - _dt.strptime(last, "%Y-%m-%d %H:%M:%S")).total_seconds()
            if elapsed < 600:
                print(f"[sync] Skipped — last sync {int(elapsed)}s ago")
                return
        except Exception:
            pass
    try:
        # Load tasks (paginated — Supabase caps single queries at 1000 rows)
        tasks = [safe_task(t) for t in _fetch_all_pages("tasks")]

        # Load teams
        res = supabase.table("teams").select("*").execute()
        teams = [r["team_name"] for r in (res.data or [])]
        if not teams:
            teams = DEFAULT_TEAMS

        # Load users (include role for role-based task permissions)
        res = supabase.table("users").select("*").execute()
        users = []
        for u in (res.data or []):
            users.append({
                "name":     u.get("name", ""),
                "team":     u.get("team", "") or "",
                "email":    u.get("email", "") or "",
                "password": u.get("password", "") or "",
                "role":     u.get("role", "") or "",
            })

        # Load initiatives (graceful: if the table is missing on this
        # account the call raises — we keep going with an empty list so
        # the rest of the app still boots).
        initiatives = []
        try:
            res = supabase.table("initiatives").select("*").execute()
            initiatives = [safe_initiative(i) for i in (res.data or [])]
        except Exception as e:
            print(f"[sync] initiatives skipped: {e}")

        # Load activity (paginated — fetch all entries, newest first)
        activity = []
        for a in _fetch_all_pages("activity_log", order_col="id", order_desc=True):
            activity.append({
                "timestamp":  a.get("timestamp", ""),
                "user":       a.get("username", ""),
                "action":     a.get("action", ""),
                "task_id":    a.get("task_id", ""),
                "task_title": a.get("task_title", ""),
                "details":    a.get("details", "")
            })

        with cache_lock:
            cache["tasks"] = tasks
            cache["teams"] = teams
            cache["users"] = users
            cache["activity"] = activity
            cache["initiatives"] = initiatives
            cache["last_sync"] = get_timestamp()

        # Rebuild FUNCTIONS so user-added functions / sub-categories show up.
        sync_functions(teams_hint=teams)

        print(f"Cache synced: {len(tasks)} tasks, {len(teams)} teams, {len(users)} users, {len(initiatives)} initiatives, {len(FUNCTIONS)} functions")

    except Exception as e:
        print(f"Cache sync failed: {e}")


# ── FUNCTIONS / SUB-CATEGORIES (user-extendable) ────────────────
def _load_sub_categories():
    """Read the sub_categories overlay from Supabase. Returns a
    {function: [sub, ...]} dict, or None if the table isn't there yet."""
    try:
        res = supabase.table("sub_categories").select("*").execute()
        out = {}
        for row in (res.data or []):
            fn = (row.get("function") or "").strip()
            name = (row.get("sub_category") or row.get("name") or "").strip()
            if fn and name:
                out.setdefault(fn, []).append(name)
        return out
    except Exception as e:
        print(f"sub_categories table not available (using built-in defaults only): {e}")
        return None


def sync_functions(teams_hint=None):
    """Rebuild the FUNCTIONS dict from teams + sub_categories tables.
    Built-in defaults are still used for any team that has no overlay row,
    so the existing seeded functions keep their sub-categories without
    requiring rows in the sub_categories table."""
    try:
        if teams_hint is None:
            tres = supabase.table("teams").select("team_name").execute()
            teams = [r["team_name"] for r in (tres.data or [])]
        else:
            teams = list(teams_hint)
        if not teams:
            teams = list(DEFAULT_FUNCTIONS.keys())

        new_fn = {}
        for fn in teams:
            # Start from built-in defaults if we know this function, else empty
            new_fn[fn] = list(DEFAULT_FUNCTIONS.get(fn, []))

        # Overlay any user-added sub-categories
        overlay = _load_sub_categories()
        if overlay:
            for fn, subs in overlay.items():
                bucket = new_fn.setdefault(fn, [])
                for s in subs:
                    if s not in bucket:
                        bucket.append(s)

        FUNCTIONS.clear()
        FUNCTIONS.update(new_fn)
        with cache_lock:
            cache["functions"] = FUNCTIONS
    except Exception as e:
        print(f"sync_functions failed (keeping previous): {e}")

def background_sync():
    while True:
        time.sleep(600)
        sync_cache_from_supabase()

# ── SEED DEFAULT TEAMS ──────────────────────────────────────────
def seed_teams():
    try:
        res = supabase.table("teams").select("team_name").execute()
        existing = [r["team_name"] for r in (res.data or [])]
        new_teams = [{"team_name": t} for t in DEFAULT_TEAMS if t not in existing]
        if new_teams:
            supabase.table("teams").insert(new_teams).execute()
            print(f"Seeded teams: {[t['team_name'] for t in new_teams]}")
            # Sheets mirror — one backup per seeded team.
            for t in new_teams:
                backup_team_to_sheets("insert", t["team_name"])
    except Exception as e:
        print(f"Team seeding failed: {e}")

# ── SERVE FRONTEND ──────────────────────────────────────────────
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@app.route("/")
def serve_frontend():
    return send_from_directory(_BASE_DIR, "index.html")

@app.route("/api/version", methods=["GET"])
def get_version():
    return jsonify({"version": APP_VERSION})

# ── API: TEAMS ──────────────────────────────────────────────────
@app.route("/api/teams", methods=["GET"])
def get_teams():
    with cache_lock:
        teams = list(cache["teams"])
    return jsonify({"teams": teams}), 200

# ── API: USERS ──────────────────────────────────────────────────
@app.route("/api/users", methods=["GET"])
def get_users():
    with cache_lock:
        users = list(cache["users"])
    safe = [{"name": u["name"], "team": u["team"], "email": u.get("email", "")} for u in users]
    return jsonify({"users": safe}), 200

# ── API: LOGIN ──────────────────────────────────────────────────
def get_user_role(username):
    """Return the lowercase role for a username, or '' if unknown.
    Falls back to a direct Supabase query on cache miss so role checks
    work across PA workers that haven't synced yet."""
    if not username:
        return ""
    uname_lower = username.lower()
    with cache_lock:
        u = next((x for x in cache["users"] if x.get("name", "").lower() == uname_lower), None)
    if u is not None:
        return (u.get("role") or "").lower()
    try:
        res = supabase.table("users").select("name,role").execute()
        for x in (res.data or []):
            if x.get("name", "").lower() == uname_lower:
                return (x.get("role") or "").lower()
    except Exception as e:
        print(f"get_user_role fallback failed: {e}")
    return ""


def can_modify_task(username, task):
    """True if `username` is admin, or is among the task's assignees."""
    if not username:
        return False
    if get_user_role(username) == "admin":
        return True
    assignees = (task.get("assigned_to") or "").lower()
    if not assignees:
        return False
    parts = [s.strip() for s in assignees.split(",") if s.strip()]
    return username.lower() in parts


def get_assigned_open_tasks(username):
    priority_order = {"high": 0, "medium": 1, "low": 2}
    uname_lower = username.lower()
    with cache_lock:
        result = [
            {k: t[k] for k in ["task_id","title","priority","status","due_date","created_by","created_at","team","sub_category"]}
            for t in cache["tasks"]
            if t["assigned_to"].lower() == uname_lower and t["status"] != "done"
        ]
    result.sort(key=lambda t: (priority_order.get(t["priority"], 3), t["due_date"] or "9999"))
    return result

@app.route("/api/login", methods=["POST"])
def login():
    try:
        data = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        if not username or not password:
            return jsonify({"error": "Username and password required."}), 400

        # 1) Look in the in-memory cache first.
        with cache_lock:
            user = next(
                (u for u in cache["users"]
                 if u["name"].lower() == username.lower() and u["password"] == password),
                None,
            )

        # 2) Fallback: cache miss can happen for users who signed up on a
        # different worker (PA runs multiple WSGI workers, each with its
        # own cache). Query Supabase directly so newly created accounts
        # can log in without waiting for the next 60s sync.
        if not user:
            try:
                res = supabase.table("users").select("*").execute()
                for u in (res.data or []):
                    if (u.get("name", "").lower() == username.lower()
                            and (u.get("password") or "") == password):
                        user = {
                            "name":     u.get("name", ""),
                            "team":     u.get("team", "") or "",
                            "email":    u.get("email", "") or "",
                            "password": u.get("password", "") or "",
                            "role":     u.get("role", "") or "",
                        }
                        with cache_lock:
                            # Add to cache so subsequent requests don't re-query.
                            if not any(c["name"].lower() == user["name"].lower()
                                       for c in cache["users"]):
                                cache["users"].append(user)
                        break
            except Exception as e:
                print(f"Login Supabase fallback failed: {e}")

        if not user:
            return jsonify({"error": "Invalid username or password."}), 401

        log_activity(user["name"], "Logged In", "", "", "")
        return jsonify({
            "message": "Login successful.",
            "user": {
                "name": user["name"],
                "team": user["team"],
                "role": user.get("role", "") or "",
            },
            "assigned_tasks": get_assigned_open_tasks(user["name"])
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: SIGNUP ──────────────────────────────────────────────────
@app.route("/api/signup", methods=["POST"])
def signup():
    try:
        data = request.get_json()
        name     = data.get("name", "").strip()
        email    = data.get("email", "").strip()
        password = data.get("password", "").strip()
        if not name or not password:
            return jsonify({"error": "Name and password are required."}), 400
        if len(password) < 4:
            return jsonify({"error": "Password must be at least 4 characters."}), 400
        with cache_lock:
            existing = next((u for u in cache["users"] if u["name"].lower() == name.lower()), None)
        if existing:
            return jsonify({"error": "Username already taken. Please log in instead."}), 409
        new_user = {"name": name, "team": "", "email": email, "password": password}

        # Synchronous write to Supabase BEFORE returning success — so the
        # row exists in the DB by the time the user tries to log in (which
        # may hit a different PA worker with its own empty cache).
        try:
            supabase.table("users").insert(new_user).execute()
        except Exception as e:
            return jsonify({"error": f"Could not create account: {e}"}), 500

        with cache_lock:
            cache["users"].append(new_user)

        # Concurrent Sheets backup — runs alongside Supabase, non-blocking.
        backup_user_to_sheets("insert", new_user)
        log_activity(name, "Signed Up", "", "", f"Email: {email}")
        return jsonify({"message": "Account created successfully.", "user": {"name": name, "team": ""}}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: FORGOT PASSWORD ─────────────────────────────────────────
# Email-based password reset. No token/email flow yet — the caller
# provides the new password directly in the body. We update Supabase
# synchronously (like signup) so the next login on any PA worker
# sees the new password regardless of which worker handled this reset.
@app.route("/api/forgot-password", methods=["POST"])
def forgot_password():
    try:
        data = request.get_json(force=True) or {}
        email        = (data.get("email") or "").strip()
        new_password = (data.get("new_password") or "").strip()

        if not email:
            return jsonify({"error": "Email is required."}), 400
        if not new_password:
            return jsonify({"error": "New password is required."}), 400
        if len(new_password) < 4:
            return jsonify({"error": "Password must be at least 4 characters."}), 400

        email_lower = email.lower()
        with cache_lock:
            user = next(
                (u for u in cache["users"]
                 if (u.get("email") or "").strip().lower() == email_lower),
                None,
            )
        if not user:
            return jsonify({"error": "No account found with this email"}), 404

        # Synchronous Supabase write — same pattern as signup so a
        # subsequent login on any worker reads the fresh password.
        try:
            supabase.table("users").update({"password": new_password}).eq(
                "name", user["name"]
            ).execute()
        except Exception as e:
            return jsonify({"error": f"Could not reset password: {e}"}), 500

        with cache_lock:
            user["password"] = new_password

        # Sheets backup: full user row update so the mirrored sheet stays honest.
        backup_user_to_sheets("update", user)
        log_activity(user.get("name", "System"),
                     "Password Reset", "", "",
                     f"Reset via email: {email}")

        return jsonify({"message": "Password updated successfully.",
                        "user": {"name": user.get("name", "")}}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: MY TASKS ────────────────────────────────────────────────
@app.route("/api/my-tasks", methods=["GET"])
def my_tasks():
    try:
        username = request.args.get("user", "").strip()
        if not username:
            return jsonify({"tasks": []}), 200
        return jsonify({"tasks": get_assigned_open_tasks(username)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: FUNCTIONS ───────────────────────────────────────────────
@app.route("/api/functions", methods=["GET"])
def get_functions():
    return jsonify({"functions": FUNCTIONS}), 200


@app.route("/api/functions/add", methods=["POST"])
def add_function():
    """Add a new top-level function (stored as a team)."""
    try:
        data = request.get_json(force=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Function name is required."}), 400
        if len(name) > 60:
            return jsonify({"error": "Function name is too long (max 60)."}), 400

        # Case-insensitive duplicate check against the current cache
        existing = {t.lower() for t in cache.get("teams", [])}
        if name.lower() in existing:
            return jsonify({"error": f"'{name}' already exists."}), 409

        try:
            supabase.table("teams").insert({"team_name": name}).execute()
        except Exception as e:
            return jsonify({"error": f"Could not save function: {e}"}), 500

        with cache_lock:
            if name not in cache["teams"]:
                cache["teams"].append(name)
        sync_functions(teams_hint=cache["teams"])

        backup_team_to_sheets("insert", name)
        log_activity(data.get("user", "System"), "Function Added", "", "", f"Function: {name}")
        return jsonify({"success": True, "function": name, "functions": FUNCTIONS}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/sub-categories/add", methods=["POST"])
def add_sub_category():
    """Add a new sub-category under an existing function."""
    try:
        data = request.get_json(force=True) or {}
        function = (data.get("function") or "").strip()
        sub_category = (data.get("sub_category") or data.get("name") or "").strip()

        if not function:
            return jsonify({"error": "Function is required."}), 400
        if not sub_category:
            return jsonify({"error": "Sub-category name is required."}), 400
        if len(sub_category) > 80:
            return jsonify({"error": "Sub-category name is too long (max 80)."}), 400
        if function not in FUNCTIONS and function not in cache.get("teams", []):
            return jsonify({"error": f"Unknown function '{function}'. Add the function first."}), 400

        # Duplicate guard (case-insensitive within the function)
        existing = {s.lower() for s in FUNCTIONS.get(function, [])}
        if sub_category.lower() in existing:
            return jsonify({"error": f"'{sub_category}' already exists under {function}."}), 409

        try:
            supabase.table("sub_categories").insert({
                "function": function,
                "sub_category": sub_category,
            }).execute()
        except Exception as e:
            msg = str(e)
            hint = ""
            lower = msg.lower()
            if "sub_categories" in lower and (
                "not find" in lower or "relation" in lower or "does not exist" in lower or "schema cache" in lower
            ):
                hint = (
                    " — Create this table in Supabase first (SQL editor): "
                    "CREATE TABLE sub_categories (id bigserial PRIMARY KEY, "
                    "function text NOT NULL, sub_category text NOT NULL, "
                    "created_at timestamptz DEFAULT now(), "
                    "UNIQUE(function, sub_category));"
                )
            return jsonify({"error": f"Could not save sub-category: {msg}{hint}"}), 500

        # Update FUNCTIONS in place so the response reflects the new state
        FUNCTIONS.setdefault(function, []).append(sub_category)
        with cache_lock:
            cache["functions"] = FUNCTIONS

        # Sheets backup — we don't have the Supabase-generated ID here so
        # the backup uses a synthetic composite key ("function::sub_category").
        # If you later query the numeric id back and want to update the
        # sheet row, call backup_subcategory_to_sheets("update", {…}) with
        # the real id.
        backup_subcategory_to_sheets("insert", {
            "function": function,
            "sub_category": sub_category,
            "created_at": get_timestamp(),
        })
        log_activity(data.get("user", "System"), "Sub-category Added", "", "",
                     f"{function} / {sub_category}")
        return jsonify({
            "success": True,
            "function": function,
            "sub_category": sub_category,
            "functions": FUNCTIONS,
        }), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: STATS ───────────────────────────────────────────────────
@app.route("/api/stats", methods=["GET"])
def get_stats():
    try:
        with cache_lock:
            all_tasks = list(cache["tasks"])
            teams = list(cache["teams"])

        today = datetime.now().strftime("%Y-%m-%d")
        filtered = all_tasks
        total = len(filtered)
        done = sum(1 for t in filtered if t["status"] == "done")
        in_progress = sum(1 for t in filtered if t["status"] in ["progress", "just"])
        blocked = sum(1 for t in filtered if t["status"] in ["blocked", "hold"])
        unassigned = sum(1 for t in filtered if not t["assigned_to"] or t["assigned_to"] == "Unassigned")
        overdue = sum(1 for t in filtered if t["status"] != "done" and t["due_date"] and t["due_date"] < today)
        complete_pct = round((done / total * 100) if total > 0 else 0)

        by_function = []
        for team in teams:
            tt = [t for t in filtered if t["team"] == team]
            if tt:
                by_function.append({"team": team, "total": len(tt), "completed": sum(1 for t in tt if t["status"] == "done")})

        by_quadrant = {q: sum(1 for t in filtered if t["quadrant"] == q and t["status"] != "done") for q in VALID_QUADRANTS}
        by_status = [{"id": s, "count": sum(1 for t in filtered if t["status"] == s)} for s in VALID_STATUSES if sum(1 for t in filtered if t["status"] == s) > 0]

        return jsonify({
            "total": total, "done": done, "in_progress": in_progress,
            "blocked": blocked, "overdue": overdue, "unassigned": unassigned,
            "complete_pct": complete_pct, "by_function": by_function,
            "by_quadrant": by_quadrant, "by_status": by_status
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: GET TASKS ───────────────────────────────────────────────
@app.route("/api/tasks", methods=["GET"])
def get_tasks():
    try:
        with cache_lock:
            tasks = list(cache["tasks"])
        team = request.args.get("team", "")
        sub = request.args.get("sub", "")
        status = request.args.get("status", "")
        priority = request.args.get("priority", "")
        quadrant = request.args.get("quadrant", "")
        owner = request.args.get("owner", "")
        work_type = request.args.get("work_type", "")
        if team and team != "All": tasks = [t for t in tasks if t["team"] == team]
        if sub and sub != "all": tasks = [t for t in tasks if t["sub_category"] == sub]
        if status and status != "all": tasks = [t for t in tasks if t["status"] == status]
        if priority and priority != "all": tasks = [t for t in tasks if t["priority"] == priority]
        if quadrant and quadrant != "all": tasks = [t for t in tasks if t["quadrant"] == quadrant]
        if owner and owner != "all": tasks = [t for t in tasks if t["assigned_to"] == owner]
        # work_type filter: exact match; 'all' or empty means no filter.
        if work_type and work_type != "all":
            tasks = [t for t in tasks if (t.get("work_type") or "") == work_type]
        return jsonify({"tasks": tasks}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: ADD SINGLE TASK ─────────────────────────────────────────
@app.route("/api/tasks", methods=["POST"])
def add_task():
    try:
        data = request.get_json()
        if not data.get("title", "").strip():
            return jsonify({"error": "'title' is required."}), 400

        priority = data.get("priority", "medium")
        if priority not in VALID_PRIORITIES: priority = "medium"
        status = data.get("status", "todo")
        if status not in VALID_STATUSES: status = "todo"
        quadrant = data.get("quadrant", "q2")
        if quadrant not in VALID_QUADRANTS: quadrant = "q2"
        team = data.get("team", "").strip()
        sub_category = data.get("sub_category", "").strip()
        if team and team in FUNCTIONS and not sub_category:
            _subs = FUNCTIONS.get(team) or []
            if _subs:
                sub_category = _subs[0]

        task_id = generate_task_id()
        now = get_timestamp()
        # If the creator left assigned_to blank (or "Unassigned"), default
        # the assignment to themselves — almost always what they wanted.
        assigned_to = (data.get("assigned_to") or "").strip()
        if not assigned_to or assigned_to.lower() == "unassigned":
            assigned_to = (data.get("user") or "").strip()
        env = (data.get("env") or "").strip().lower()
        if env not in VALID_ENVS:
            env = ""
        created_by = data.get("user", "System")
        new_task = safe_task({
            "task_id": task_id,
            "title": data["title"].strip(),
            "description": data.get("description", "").strip(),
            "assigned_to": assigned_to,
            "team": team,
            "due_date": normalize_date(data.get("due_date", "")),
            "priority": priority,
            "status": status,
            "created_by": created_by,
            "created_at": now,
            "updated_at": now,
            "sub_category": sub_category,
            "quadrant": quadrant,
            "notes": data.get("notes", "").strip(),
            "start_date": normalize_date(data.get("start_date", "")),
            "attachments": data.get("attachments", []),
            "parent_task_id": (data.get("parent_task_id") or "").strip(),
            "env": env,
            "depends_on": [str(x).strip() for x in (data.get("depends_on") or []) if str(x).strip()],
            "milestone_marker": bool(data.get("milestone_marker") or False),
            "initiative_id": (data.get("initiative_id") or "").strip(),
            "work_type": ((data.get("work_type") or "").strip().lower()
                          if (data.get("work_type") or "").strip().lower() in VALID_WORK_TYPES
                          else ""),
            "status_log": [{"status": status, "changed_at": now, "changed_by": created_by}],
        })

        with cache_lock:
            cache["tasks"].append(new_task)

        def db_write():
            supabase.table("tasks").insert(new_task).execute()
        write_in_background(db_write)
        # Sheets backup fans out on its own daemon thread — runs in parallel with Supabase.
        backup_task_to_sheets("insert", new_task)
        log_activity(data.get("user", "System"), "Created", task_id, new_task["title"], f"Function: {team}, Priority: {priority}")
        return jsonify({"message": "Task created successfully.", "task": new_task}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: BULK ADD TASKS ──────────────────────────────────────────
@app.route("/api/tasks/bulk", methods=["POST"])
def bulk_add_tasks():
    try:
        data = request.get_json()
        titles = [t.strip() for t in data.get("titles", []) if t.strip()]
        if not titles:
            return jsonify({"error": "No task titles provided."}), 400

        priority = data.get("priority", "medium")
        if priority not in VALID_PRIORITIES: priority = "medium"
        status = data.get("status", "todo")
        if status not in VALID_STATUSES: status = "todo"
        quadrant = data.get("quadrant", "q2")
        if quadrant not in VALID_QUADRANTS: quadrant = "q2"
        team = data.get("team", "").strip()
        sub_category = data.get("sub_category", "").strip()
        if team and team in FUNCTIONS and not sub_category:
            _subs = FUNCTIONS.get(team) or []
            if _subs:
                sub_category = _subs[0]
        assigned_to = (data.get("assigned_to") or "").strip()
        if not assigned_to or assigned_to.lower() == "unassigned":
            assigned_to = (data.get("user") or "").strip()
        due_date = normalize_date(data.get("due_date", ""))
        start_date = normalize_date(data.get("start_date", ""))
        work_type = (data.get("work_type") or "").strip().lower()
        if work_type not in VALID_WORK_TYPES:
            work_type = ""
        now = get_timestamp()

        new_tasks = []
        for title in titles:
            task_id = generate_task_id()
            new_tasks.append(safe_task({
                "task_id": task_id, "title": title, "description": "",
                "assigned_to": assigned_to, "team": team, "due_date": due_date,
                "priority": priority, "status": status,
                "created_by": data.get("user", "System"),
                "created_at": now, "updated_at": now,
                "sub_category": sub_category, "quadrant": quadrant,
                "notes": "", "start_date": start_date, "attachments": [],
                "work_type": work_type,
            }))

        with cache_lock:
            cache["tasks"].extend(new_tasks)

        def db_write():
            supabase.table("tasks").insert(new_tasks).execute()
        write_in_background(db_write)
        backup_tasks_to_sheets("insert", new_tasks)
        log_activity(data.get("user", "System"), "Bulk Created", "", f"{len(new_tasks)} tasks", f"Function: {team}")
        return jsonify({"message": f"{len(new_tasks)} tasks created.", "tasks": new_tasks, "count": len(new_tasks)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: IMPORT TASKS FROM CSV OR XLSX ──────────────────────────
def _normalize_header(h):
    """Lowercase, trim, drop parenthetical hints like '(YYYY-MM-DD)',
    and replace spaces/dashes with underscores so the template header
    'due_date (YYYY-MM-DD)' still maps to 'due_date'."""
    s = (h or "").strip()
    if "(" in s:
        s = s.split("(", 1)[0]
    return s.strip().lower().replace(" ", "_").replace("-", "_")


def _coerce_cell(v):
    """Coerce a cell value (CSV string, XLSX number, XLSX date, etc.) to
    a plain string. Real date/datetime objects become ISO YYYY-MM-DD."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):           # datetime.date / datetime.datetime
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_csv_rows(raw):
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(content))
    return list(reader.fieldnames or []), [dict(r) for r in reader]


def _parse_xlsx_rows(raw):
    try:
        import openpyxl  # noqa - imported lazily so .csv-only deploys keep working
    except ImportError:
        raise RuntimeError("openpyxl is required to import .xlsx files. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in all_rows[0]]
    out = []
    for row in all_rows[1:]:
        if all(c in (None, "") for c in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if h:
                d[h] = _coerce_cell(v)
        out.append(d)
    return headers, out


@app.route("/api/tasks/import-csv", methods=["POST"])
def import_tasks_csv():
    """Bulk-create tasks from an uploaded .csv OR .xlsx file.
    Required column: title. All other columns are optional and fall
    back to the same defaults the regular /api/tasks endpoint uses.
    Date columns may include format hints like 'due_date (YYYY-MM-DD)'."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded (expected multipart field 'file')."}), 400
        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"error": "Empty filename."}), 400

        fname = f.filename.lower()
        raw = f.read()
        try:
            if fname.endswith(".xlsx"):
                fieldnames, rows = _parse_xlsx_rows(raw)
            elif fname.endswith(".csv"):
                fieldnames, rows = _parse_csv_rows(raw)
            else:
                return jsonify({"error": "Only .csv and .xlsx files are accepted."}), 400
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 500
        except Exception as e:
            return jsonify({"error": f"Could not parse file: {e}"}), 400

        if not fieldnames:
            return jsonify({"error": "File is empty or has no header row."}), 400

        # Case-insensitive header mapping (also strips format hints in parens).
        field_map = {}
        for fn in fieldnames:
            key = _normalize_header(fn)
            if key:
                field_map[key] = fn
        if "title" not in field_map:
            return jsonify({
                "error": "File must have a 'title' column.",
                "found_columns": list(fieldnames),
            }), 400

        def cell(row, name):
            src = field_map.get(name)
            if not src:
                return ""
            v = row.get(src, "")
            return (v or "").strip() if isinstance(v, str) else _coerce_cell(v)

        now = get_timestamp()
        user = (request.form.get("user") or "System").strip() or "System"

        created = []
        skipped_no_title = 0
        errors = []

        for i, row in enumerate(rows, start=2):  # row 2 = first data row
            title = cell(row, "title")
            if not title:
                skipped_no_title += 1
                continue
            try:
                priority = (cell(row, "priority") or "medium").lower()
                if priority not in VALID_PRIORITIES:
                    priority = "medium"
                status = (cell(row, "status") or "todo").lower()
                if status not in VALID_STATUSES:
                    status = "todo"
                quadrant = (cell(row, "quadrant") or "q2").lower()
                if quadrant not in VALID_QUADRANTS:
                    quadrant = "q2"

                team = cell(row, "team")
                sub_category = cell(row, "sub_category")
                if team and team in FUNCTIONS and not sub_category:
                    _subs = FUNCTIONS.get(team) or []
                    if _subs:
                        sub_category = _subs[0]

                new_task = safe_task({
                    "task_id":      generate_task_id(),
                    "title":        title,
                    "description":  cell(row, "description"),
                    "assigned_to":  cell(row, "assigned_to"),
                    "team":         team,
                    "due_date":     normalize_date(cell(row, "due_date")),
                    "priority":     priority,
                    "status":       status,
                    "created_by":   cell(row, "created_by") or user,
                    "created_at":   now,
                    "updated_at":   now,
                    "sub_category": sub_category,
                    "quadrant":     quadrant,
                    "notes":        cell(row, "notes"),
                    "start_date":   normalize_date(cell(row, "start_date")),
                    "attachments":  [],
                })
                created.append(new_task)
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        if created:
            with cache_lock:
                cache["tasks"].extend(created)

            def db_write(rows=created):
                supabase.table("tasks").insert(rows).execute()
            write_in_background(db_write)

            log_activity(
                user, "CSV Import", "", f"{len(created)} tasks",
                f"from {f.filename}"
            )

        return jsonify({
            "success": True,
            "created": len(created),
            "skipped": skipped_no_title,
            "errors": errors[:10],
            "filename": f.filename,
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: DOWNLOAD XLSX TEMPLATE ─────────────────────────────────
@app.route("/api/tasks/template.xlsx", methods=["GET"])
def download_xlsx_template():
    """Generate and serve an .xlsx template with:
    - Date-formatted cells on due_date / start_date (Excel shows the
      native calendar picker when the user double-clicks the cell)
    - Drop-down lists on status / priority / quadrant
    - Frozen header row + bold purple header
    - One example data row so the user sees the shape immediately"""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"error": "openpyxl is not installed on this server. Run: pip install openpyxl"}), 500
        from datetime import date

        headers = [
            "title", "description", "team", "sub_category", "assigned_to",
            "priority", "status", "quadrant",
            "due_date (YYYY-MM-DD)", "start_date (YYYY-MM-DD)",
            "notes",
        ]
        DATE_COL_KEYS = {"due_date", "start_date"}
        col_index = {h: i + 1 for i, h in enumerate(headers)}

        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        ws.append(headers)

        # NamedStyle is the most deterministic way to force a custom date
        # format: it writes the format code into xl/styles.xml so Excel
        # respects it regardless of the user's locale.
        iso_date_style = NamedStyle(name="iso_date", number_format="yyyy-mm-dd")
        wb.add_named_style(iso_date_style)

        # Style the header row
        header_fill = PatternFill("solid", fgColor="5B21B6")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell_ in ws[1]:
            cell_.fill = header_fill
            cell_.font = header_font
            cell_.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Example row — gives the user an immediate visual cue for each column
        ws.append([
            "Example task — only title is required",
            "Optional description",
            "Development",
            "Backend",
            "Alice, Bob",
            "medium",
            "todo",
            "q2",
            date(2026, 6, 30),
            date(2026, 6, 15),
            "task",
            "Any notes here",
        ])

        # Build a lookup: header-key -> column number, ignoring the
        # parenthetical hint so DATE_COL_KEYS still match.
        def _key(h):
            return h.split("(", 1)[0].strip().lower().replace(" ", "_")
        col_by_key = {_key(h): i + 1 for i, h in enumerate(headers)}

        # Set column widths + apply the NamedStyle to every date cell.
        for h, col in col_index.items():
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = max(len(h) + 4, 18)
        for k in DATE_COL_KEYS:
            col = col_by_key[k]
            letter = get_column_letter(col)
            # Apply NamedStyle to example row + all future blank rows so
            # both the displayed example and anything the user types
            # render as YYYY-MM-DD.
            for r in range(2, 1001):
                ws[f"{letter}{r}"].style = "iso_date"

        # Drop-down lists for the validated columns
        def add_dropdown(col_letter, options):
            dv = DataValidation(type="list",
                                formula1='"' + ",".join(options) + '"',
                                allow_blank=True, showDropDown=False)
            dv.error = "Please pick a value from the list."
            dv.errorTitle = "Invalid value"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        add_dropdown(get_column_letter(col_index["status"]),
                     ["todo", "just", "progress", "hold", "blocked", "review", "done", "reopen"])
        add_dropdown(get_column_letter(col_index["priority"]),
                     ["high", "medium", "low"])
        add_dropdown(get_column_letter(col_index["quadrant"]),
                     ["q1", "q2", "q3", "q4"])

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="taskflow_template.xlsx",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: EDIT TASK ───────────────────────────────────────────────
@app.route("/api/tasks/<task_id>", methods=["PUT"])
def edit_task(task_id):
    try:
        data = request.get_json()
        with cache_lock:
            current = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not current:
            return jsonify({"error": "Task not found."}), 404

        # Edit is unrestricted — any logged-in user can edit any task.
        # Delete remains admin-only (enforced in delete_task / bulk_delete_tasks).

        old_status = current.get("status")
        updated = dict(current)
        updated["updated_at"] = get_timestamp()
        updatable = ["title","description","assigned_to","team","due_date","priority","status","sub_category","quadrant","notes","start_date","attachments","parent_task_id","env","depends_on","milestone_marker","initiative_id","work_type"]
        for f in updatable:
            if f in data:
                val = data[f]
                if f in ("due_date", "start_date"):
                    updated[f] = normalize_date(val)
                elif f == "env":
                    updated[f] = (val or "").strip().lower() if isinstance(val, str) else ""
                elif f == "depends_on":
                    updated[f] = [str(x).strip() for x in (val or []) if str(x).strip()] if isinstance(val, list) else []
                elif f == "milestone_marker":
                    updated[f] = bool(val)
                elif f == "work_type":
                    v = (val or "").strip().lower() if isinstance(val, str) else ""
                    updated[f] = v if v in VALID_WORK_TYPES else current.get("work_type", "")
                else:
                    updated[f] = val.strip() if isinstance(val, str) else val

        if updated["priority"] not in VALID_PRIORITIES: updated["priority"] = "medium"
        if updated["status"] not in VALID_STATUSES: updated["status"] = "todo"
        if updated["quadrant"] not in VALID_QUADRANTS: updated["quadrant"] = "q2"
        if updated.get("env") not in VALID_ENVS: updated["env"] = ""

        if updated["status"] != old_status:
            entry = {"status": updated["status"], "changed_at": updated["updated_at"], "changed_by": data.get("user", "System")}
            updated["status_log"] = list(current.get("status_log") or []) + [entry]

        # Auto-assign tester when a dev-function task moves to done.
        _tester_assigned = None
        if updated["status"] == "done" and old_status != "done":
            new_assignees, _tester_assigned = _auto_assign_tester(updated)
            if new_assignees is not None:
                updated["assigned_to"] = new_assignees

        with cache_lock:
            cache["tasks"] = [updated if t["task_id"] == task_id else t for t in cache["tasks"]]

        def db_write():
            supabase.table("tasks").update(updated).eq("task_id", task_id).execute()
        write_in_background(db_write)
        backup_task_to_sheets("update", updated)
        def _val_str(v):
            if isinstance(v, list): return ", ".join(str(x) for x in v) if v else ""
            return str(v or "")
        def _changed(old, new):
            if isinstance(new, list) or isinstance(old, list):
                return sorted(str(x) for x in (new or [])) != sorted(str(x) for x in (old or []))
            return str(new or "") != str(old or "")
        changes = [
            f"{f}: {_val_str(current.get(f))} → {_val_str(updated[f])}"
            for f in updatable
            if f in data and _changed(current.get(f), updated[f])
        ]
        if changes:
            log_activity(data.get("user", "System"), "Edited", task_id, updated["title"], ", ".join(changes))
        if _tester_assigned:
            log_activity(data.get("user", "System"), "Tester Auto-Assigned", task_id, updated["title"],
                         f"Auto-assigned tester: {_tester_assigned} (status changed to done)")
        return jsonify({"message": "Task updated successfully.", "task": updated}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: TOGGLE STATUS ───────────────────────────────────────────
@app.route("/api/tasks/<task_id>/toggle", methods=["PATCH"])
def toggle_task(task_id):
    try:
        data = request.get_json(silent=True) or {}
        req_user = data.get("user", "System")
        with cache_lock:
            current = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not current:
            return jsonify({"error": "Task not found."}), 404
        new_status = "done" if current["status"] != "done" else "todo"
        now = get_timestamp()
        new_status_log = list(current.get("status_log") or []) + [{"status": new_status, "changed_at": now, "changed_by": req_user}]

        # Auto-assign tester when toggling to done.
        new_assigned_to = current.get("assigned_to") or ""
        _tester_assigned = None
        if new_status == "done":
            _probe = dict(current)
            _probe["status"] = "done"
            new_assignees, _tester_assigned = _auto_assign_tester(_probe)
            if new_assignees is not None:
                new_assigned_to = new_assignees

        with cache_lock:
            for t in cache["tasks"]:
                if t["task_id"] == task_id:
                    t["status"] = new_status
                    t["updated_at"] = now
                    t["status_log"] = new_status_log
                    if _tester_assigned:
                        t["assigned_to"] = new_assigned_to

        db_payload = {"status": new_status, "updated_at": now, "status_log": new_status_log}
        if _tester_assigned:
            db_payload["assigned_to"] = new_assigned_to

        def db_write():
            supabase.table("tasks").update(db_payload).eq("task_id", task_id).execute()
        write_in_background(db_write)
        log_activity(req_user, "Status Changed", task_id, current["title"], f"status: {current['status']} → {new_status}")
        if _tester_assigned:
            log_activity(req_user, "Tester Auto-Assigned", task_id, current["title"],
                         f"Auto-assigned tester: {_tester_assigned} (status changed to done)")
        return jsonify({"message": f"Task marked as {new_status}.", "task_id": task_id, "status": new_status,
                        "assigned_to": new_assigned_to}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: DUPLICATE TASK ──────────────────────────────────────────
@app.route("/api/tasks/<task_id>/duplicate", methods=["POST"])
def duplicate_task(task_id):
    try:
        data = request.get_json(silent=True) or {}
        req_user = data.get("user", "System")
        with cache_lock:
            original = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not original:
            return jsonify({"error": "Task not found."}), 404
        now = get_timestamp()
        new_task = dict(original)
        new_task["task_id"] = generate_task_id()
        new_task["title"] = f"{original['title']} (copy)"
        new_task["status"] = "todo"
        new_task["created_at"] = now
        new_task["updated_at"] = now
        with cache_lock:
            cache["tasks"].insert(0, new_task)
        def db_write():
            supabase.table("tasks").insert(new_task).execute()
        write_in_background(db_write)
        log_activity(req_user, "Duplicated", new_task["task_id"], new_task["title"], f"From: {task_id}")
        return jsonify({"message": "Task duplicated.", "task": new_task}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: DELETE TASK ─────────────────────────────────────────────
@app.route("/api/tasks/<task_id>", methods=["DELETE"])
def delete_task(task_id):
    try:
        req_user = request.args.get("user", "System")
        with cache_lock:
            deleted = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not deleted:
            return jsonify({"error": "Task not found."}), 404

        # Delete is admin-only.
        if get_user_role(req_user) != "admin":
            return jsonify({"error": "Forbidden: only admins can delete tasks."}), 403

        with cache_lock:
            cache["tasks"] = [t for t in cache["tasks"] if t["task_id"] != task_id]
        def db_write():
            supabase.table("tasks").delete().eq("task_id", task_id).execute()
        write_in_background(db_write)
        backup_task_to_sheets("delete", {"task_id": task_id})
        log_activity(req_user, "Deleted", task_id, deleted.get("title", ""), "")
        return jsonify({"message": "Task deleted.", "task_id": task_id}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: BULK UPDATE ─────────────────────────────────────────────
@app.route("/api/tasks/bulk-update", methods=["PATCH"])
def bulk_update_tasks():
    try:
        data = request.get_json()
        task_ids = data.get("task_ids", [])
        field = data.get("field", "").strip()
        value = data.get("value", "").strip()
        if not task_ids or not field:
            return jsonify({"error": "task_ids and field are required."}), 400
        allowed = ["status","priority","quadrant","assigned_to","team","sub_category"]
        if field not in allowed:
            return jsonify({"error": f"Field must be one of: {', '.join(allowed)}"}), 400
        now = get_timestamp()
        # Snapshot the mutated task dicts (as they'll appear post-update)
        # so the Sheets backup writes the same rows Supabase will see.
        updated_snapshot = []
        # task_id → tester_name for tasks that got a tester auto-appended.
        tester_assignments = {}
        with cache_lock:
            for t in cache["tasks"]:
                if t["task_id"] in task_ids:
                    t[field] = value
                    t["updated_at"] = now
                    if field == "status" and value == "done":
                        new_assignees, tester_name = _auto_assign_tester(t)
                        if new_assignees is not None:
                            t["assigned_to"] = new_assignees
                            tester_assignments[t["task_id"]] = tester_name
                    updated_snapshot.append(dict(t))

        def db_write(snap=updated_snapshot, ta=tester_assignments):
            supabase.table("tasks").update({field: value, "updated_at": now}).in_("task_id", task_ids).execute()
            # Patch assigned_to individually for tasks that got a tester appended.
            for tid, _ in ta.items():
                task_snap = next((t for t in snap if t["task_id"] == tid), None)
                if task_snap:
                    supabase.table("tasks").update({"assigned_to": task_snap["assigned_to"]}).eq("task_id", tid).execute()
        write_in_background(db_write)
        backup_tasks_to_sheets("update", updated_snapshot)
        log_activity(data.get("user", "System"), "Bulk Updated", "", f"{len(task_ids)} tasks", f"Set {field} = {value}")
        if tester_assignments:
            tester_names = ", ".join(sorted(set(tester_assignments.values())))
            log_activity(data.get("user", "System"), "Tester Auto-Assigned", "",
                         f"{len(tester_assignments)} task(s)",
                         f"Auto-assigned tester: {tester_names} (status changed to done)")
        return jsonify({"message": f"Updated {len(task_ids)} tasks.", "count": len(task_ids)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: BULK DELETE ─────────────────────────────────────────────
@app.route("/api/tasks/bulk-delete", methods=["POST"])
def bulk_delete_tasks():
    try:
        data = request.get_json()
        task_ids = data.get("task_ids", [])
        if not task_ids:
            return jsonify({"error": "task_ids is required."}), 400

        # Bulk delete is admin-only.
        req_user = (data.get("user") or "").strip()
        if get_user_role(req_user) != "admin":
            return jsonify({"error": "Forbidden: only admins can delete tasks."}), 403

        with cache_lock:
            cache["tasks"] = [t for t in cache["tasks"] if t["task_id"] not in task_ids]
        def db_write():
            supabase.table("tasks").delete().in_("task_id", task_ids).execute()
        write_in_background(db_write)
        # Sheets backup expects task dicts; we only need the task_id for delete.
        backup_tasks_to_sheets("delete", [{"task_id": tid} for tid in task_ids])
        log_activity(req_user or "System", "Bulk Deleted", "", f"{len(task_ids)} tasks", "")
        return jsonify({"message": f"Deleted {len(task_ids)} tasks.", "count": len(task_ids)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: TASK ATTACHMENTS (Supabase Storage) ─────────────────────
# Bucket is public so we hand back a permanent public URL; no signed-URL dance.
_ATTACHMENT_BUCKET   = "taskflow-attachments"
_ATTACHMENT_ALLOWED  = {"jpg", "jpeg", "png", "gif", "mp4", "mov", "webm"}


def _normalize_attachments_field(value):
    """The cache may hold a list (new format) or a legacy semicolon/comma
    string. Always return a clean list of URLs."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    s = str(value).strip()
    if not s:
        return []
    parts = s.replace(";", ",").split(",")
    return [p.strip() for p in parts if p.strip()]


@app.route("/api/tasks/<task_id>/attachments", methods=["POST"])
def upload_task_attachment(task_id):
    """Upload one file to Supabase Storage and append its public URL to
    the task's `attachments` JSONB array. Whitelisted: images + videos."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file in request (multipart field 'file')."}), 400
        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"error": "Empty filename."}), 400

        ext = os.path.splitext(f.filename)[1].lower().lstrip(".")
        if ext not in _ATTACHMENT_ALLOWED:
            return jsonify({
                "error": f".{ext} is not allowed. Allowed: {', '.join(sorted(_ATTACHMENT_ALLOWED))}"
            }), 400

        with cache_lock:
            task = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not task:
            return jsonify({"error": "Task not found."}), 404

        safe_base = "".join(c if c.isalnum() or c in "._-" else "_" for c in f.filename)
        uid = uuid.uuid4().hex[:10]
        storage_path = f"{task_id}/{uid}_{safe_base}"

        file_data = f.read()
        content_type = f.content_type or "application/octet-stream"

        try:
            supabase.storage.from_(_ATTACHMENT_BUCKET).upload(
                storage_path,
                file_data,
                file_options={"content-type": content_type},
            )
        except Exception as e:
            return jsonify({"error": f"Storage upload failed: {e}"}), 500

        # Build public URL directly — avoids supabase-py version differences
        # (some return dict, some str from get_public_url).
        public_url = f"{SUPABASE_URL}/storage/v1/object/public/{_ATTACHMENT_BUCKET}/{storage_path}"

        new_attachments = _normalize_attachments_field(task.get("attachments")) + [public_url]

        try:
            supabase.table("tasks").update({"attachments": new_attachments}).eq("task_id", task_id).execute()
        except Exception as e:
            return jsonify({"error": f"Could not update task row: {e}"}), 500

        with cache_lock:
            task["attachments"] = new_attachments
            task["updated_at"]  = get_timestamp()

        log_activity(request.form.get("user", "System"),
                     "Attachment Added", task_id,
                     task.get("title", ""), f.filename)

        return jsonify({
            "attachments":   new_attachments,
            "url":           public_url,
            "filename":      storage_path,
            "original_name": f.filename,
            "size":          len(file_data),
            "type":          ("image" if content_type.startswith("image")
                              else "video" if content_type.startswith("video")
                              else "file"),
        }), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tasks/<task_id>/attachments", methods=["DELETE"])
def delete_task_attachment(task_id):
    """Remove an attachment from Supabase Storage and from the task's
    `attachments` array. Body: {filename, ...} OR {url, ...} — either."""
    try:
        data = request.get_json(force=True) or {}
        filename = (data.get("filename") or "").strip()
        url      = (data.get("url") or "").strip()
        # Derive storage path from URL if needed
        if not filename and url:
            sep = f"/{_ATTACHMENT_BUCKET}/"
            if sep in url:
                filename = url.split(sep, 1)[1]
        if not filename:
            return jsonify({"error": "filename or url is required in the body."}), 400

        with cache_lock:
            task = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not task:
            return jsonify({"error": "Task not found."}), 404

        # Best-effort storage delete — don't fail the DB update if the
        # object's already gone or storage is having a bad day.
        try:
            supabase.storage.from_(_ATTACHMENT_BUCKET).remove([filename])
        except Exception as e:
            print(f"[attachments DELETE] storage warning: {e}")

        current = _normalize_attachments_field(task.get("attachments"))
        new_attachments = [u for u in current if filename not in u]

        try:
            supabase.table("tasks").update({"attachments": new_attachments}).eq("task_id", task_id).execute()
        except Exception as e:
            return jsonify({"error": f"Could not update task row: {e}"}), 500

        with cache_lock:
            task["attachments"] = new_attachments
            task["updated_at"]  = get_timestamp()

        log_activity(data.get("user", "System"),
                     "Attachment Removed", task_id,
                     task.get("title", ""), filename)

        return jsonify({"attachments": new_attachments, "removed": filename}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: TASK COMMENTS ───────────────────────────────────────────
# Append a comment to a task's JSONB `comments` array. Each comment
# is {text, by, at}. The Supabase `tasks` table needs a `comments`
# jsonb column for persistence; if it's missing the row update will
# fail and we surface that error verbatim so it's clear what to add.
@app.route("/api/tasks/<task_id>/comments", methods=["POST"])
def add_task_comment(task_id):
    try:
        data = request.get_json(force=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "Comment text is required."}), 400

        with cache_lock:
            task = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not task:
            return jsonify({"error": "Task not found."}), 404

        comment = {
            "text": text,
            "by":   (data.get("user") or "System").strip() or "System",
            "at":   get_timestamp(),
        }
        comments = list(task.get("comments") or [])
        comments.append(comment)

        try:
            supabase.table("tasks").update({"comments": comments}).eq("task_id", task_id).execute()
        except Exception as e:
            # If the Supabase column doesn't exist yet, the in-memory
            # cache still gets the comment — but warn the caller.
            print(f"[comments] Supabase update failed: {e}")
            return jsonify({
                "error": f"Saved to cache only; Supabase update failed: {e}. "
                         f"Add a JSONB `comments` column to the tasks table."
            }), 500

        with cache_lock:
            task["comments"]   = comments
            task["updated_at"] = get_timestamp()

        log_activity(comment["by"], "Comment Added", task_id,
                     task.get("title", ""), text[:80])

        return jsonify({"comments": comments, "comment": comment}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: CLIENT CONFIG (for direct browser → Supabase uploads) ───
# Hands the browser the anon key + bucket info so it can POST files
# straight to Supabase Storage and skip the Flask hop. The "anon" key
# is *meant* for client-side use (Supabase RLS gates what it can do);
# it's the same key the server already uses to talk to the bucket.
@app.route("/api/config", methods=["GET"])
def get_client_config():
    try:
        return jsonify({
            "supabase_url":            SUPABASE_URL,
            "supabase_anon_key":       SUPABASE_KEY,
            "attachment_bucket":       _ATTACHMENT_BUCKET,
            "attachment_allowed_ext":  sorted(_ATTACHMENT_ALLOWED),
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: REGISTER an already-uploaded URL onto a task ────────────
# Used after a direct browser upload in edit mode — the bytes are
# already in storage, we just need to append the URL to the task row.
@app.route("/api/tasks/<task_id>/attachments/register", methods=["POST"])
def register_task_attachment(task_id):
    try:
        data = request.get_json(force=True) or {}
        url = (data.get("url") or "").strip()
        if not url:
            return jsonify({"error": "url is required"}), 400

        with cache_lock:
            task = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not task:
            return jsonify({"error": "Task not found."}), 404

        current = _normalize_attachments_field(task.get("attachments"))
        if url in current:
            # Idempotent — already attached
            return jsonify({"attachments": current, "url": url, "already": True}), 200
        new_attachments = current + [url]

        try:
            supabase.table("tasks").update({"attachments": new_attachments}).eq("task_id", task_id).execute()
        except Exception as e:
            return jsonify({"error": f"Could not update task row: {e}"}), 500

        with cache_lock:
            task["attachments"] = new_attachments
            task["updated_at"]  = get_timestamp()

        log_activity(data.get("user", "System"),
                     "Attachment Added", task_id,
                     task.get("title", ""), url.rsplit("/", 1)[-1])

        return jsonify({"attachments": new_attachments, "url": url}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: STANDALONE ATTACHMENT UPLOAD (for new tasks not yet saved) ──
# Same bucket / same whitelist as the task-bound endpoints; files land at
# `_new/<uid>_<safename>` and the caller gets back a public URL. When the
# task is later POSTed, the URL is included in its `attachments` array.
# Files never claimed by a save become orphans in the public bucket —
# harmless, and the bucket can be swept periodically if it ever matters.
@app.route("/api/attachments", methods=["POST"])
def upload_attachment_standalone():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file in request (multipart field 'file')."}), 400
        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"error": "Empty filename."}), 400

        ext = os.path.splitext(f.filename)[1].lower().lstrip(".")
        if ext not in _ATTACHMENT_ALLOWED:
            return jsonify({
                "error": f".{ext} is not allowed. Allowed: {', '.join(sorted(_ATTACHMENT_ALLOWED))}"
            }), 400

        safe_base = "".join(c if c.isalnum() or c in "._-" else "_" for c in f.filename)
        uid = uuid.uuid4().hex[:10]
        storage_path = f"_new/{uid}_{safe_base}"

        file_data = f.read()
        content_type = f.content_type or "application/octet-stream"

        try:
            supabase.storage.from_(_ATTACHMENT_BUCKET).upload(
                storage_path,
                file_data,
                file_options={"content-type": content_type},
            )
        except Exception as e:
            return jsonify({"error": f"Storage upload failed: {e}"}), 500

        public_url = f"{SUPABASE_URL}/storage/v1/object/public/{_ATTACHMENT_BUCKET}/{storage_path}"

        return jsonify({
            "url":           public_url,
            "filename":      storage_path,
            "original_name": f.filename,
            "size":          len(file_data),
            "type":          ("image" if content_type.startswith("image")
                              else "video" if content_type.startswith("video")
                              else "file"),
        }), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/attachments", methods=["DELETE"])
def delete_attachment_standalone():
    """Best-effort remove of a single object from the bucket. Used when
    the user removes an upload from a new-task modal before saving."""
    try:
        data = request.get_json(force=True) or {}
        filename = (data.get("filename") or "").strip()
        url      = (data.get("url") or "").strip()
        if not filename and url:
            sep = f"/{_ATTACHMENT_BUCKET}/"
            if sep in url:
                filename = url.split(sep, 1)[1]
        if not filename:
            return jsonify({"error": "filename or url is required in the body."}), 400
        try:
            supabase.storage.from_(_ATTACHMENT_BUCKET).remove([filename])
        except Exception as e:
            print(f"[attachments DELETE standalone] storage warning: {e}")
        return jsonify({"removed": filename}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: INITIATIVES ─────────────────────────────────────────────
# All five endpoints read/write through the in-memory cache the same
# way the task endpoints do, with synchronous writes to Supabase so a
# subsequent call on any PA worker sees the change immediately.

@app.route("/api/initiatives", methods=["GET"])
def get_initiatives():
    try:
        with cache_lock:
            items = [dict(i) for i in cache.get("initiatives", [])]
        return jsonify({"initiatives": items, "count": len(items)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/initiatives", methods=["POST"])
def add_initiative():
    try:
        data = request.get_json(force=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "'name' is required."}), 400

        status = (data.get("status") or "planning").strip().lower()
        if status not in VALID_INITIATIVE_STATUSES:
            status = "planning"
        accent = (data.get("accent_color") or "violet").strip().lower()
        if accent not in VALID_INITIATIVE_COLORS:
            accent = "violet"
        priority = (data.get("priority") or "P2").strip().upper()
        if priority not in VALID_INITIATIVE_PRIORITIES:
            priority = "P2"

        now = get_timestamp()
        new_initiative = safe_initiative({
            "id":           generate_initiative_id(),
            "name":         name,
            "type":         (data.get("type") or "").strip(),
            "description":  (data.get("description") or "").strip(),
            "start_date":   normalize_date(data.get("start_date", "")),
            "end_date":     normalize_date(data.get("end_date", "")),
            "status":       status,
            "priority":     priority,
            "created_by":   (data.get("user") or "System").strip(),
            "accent_color": accent,
            "created_at":   now,
            "updated_at":   now,
        })

        # Synchronous write so the next GET (possibly on a different PA
        # worker) sees the new row.
        try:
            supabase.table("initiatives").insert(new_initiative).execute()
        except Exception as e:
            return jsonify({"error": f"Could not create initiative: {e}"}), 500

        with cache_lock:
            cache["initiatives"].append(new_initiative)

        log_activity(new_initiative["created_by"], "Initiative Created",
                     new_initiative["id"], name,
                     f"Status: {status}, Color: {accent}")
        return jsonify({"message": "Initiative created.",
                        "initiative": new_initiative}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/initiatives/<init_id>", methods=["GET"])
def get_initiative(init_id):
    try:
        with cache_lock:
            init = next((i for i in cache.get("initiatives", []) if i["id"] == init_id), None)
            tasks = list(cache.get("tasks", []))
        if not init:
            return jsonify({"error": "Initiative not found."}), 404

        # Computed stats — derived from current task cache so they stay
        # fresh without writing anything back to the DB.
        own = [t for t in tasks if (t.get("initiative_id") or "") == init_id]
        functions_set = {(t.get("team") or "").strip() for t in own if (t.get("team") or "").strip()}
        owners_set = set()
        for t in own:
            for n in (t.get("assigned_to") or "").split(","):
                n = n.strip()
                if n and n.lower() != "unassigned":
                    owners_set.add(n)
        done_count = sum(1 for t in own if (t.get("status") or "") == "done")

        out = dict(init)
        out["stats"] = {
            "task_count":  len(own),
            "done_count":  done_count,
            "open_count":  len(own) - done_count,
            "functions":   sorted(functions_set),
            "owners":      sorted(owners_set),
        }
        return jsonify({"initiative": out}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/initiatives/<init_id>", methods=["PUT"])
def edit_initiative(init_id):
    try:
        data = request.get_json(force=True) or {}
        with cache_lock:
            current = next((i for i in cache.get("initiatives", []) if i["id"] == init_id), None)
        if not current:
            return jsonify({"error": "Initiative not found."}), 404

        updated = dict(current)
        updated["updated_at"] = get_timestamp()
        updatable = ["name", "type", "description", "start_date", "end_date",
                     "status", "accent_color", "priority"]
        for f in updatable:
            if f in data:
                val = data[f]
                if f in ("start_date", "end_date"):
                    updated[f] = normalize_date(val)
                elif f == "status":
                    v = (val or "").strip().lower()
                    updated[f] = v if v in VALID_INITIATIVE_STATUSES else current.get("status", "planning")
                elif f == "accent_color":
                    v = (val or "").strip().lower()
                    updated[f] = v if v in VALID_INITIATIVE_COLORS else current.get("accent_color", "violet")
                elif f == "priority":
                    v = (val or "P2").strip().upper()
                    updated[f] = v if v in VALID_INITIATIVE_PRIORITIES else current.get("priority", "P2")
                else:
                    updated[f] = val.strip() if isinstance(val, str) else val
        # Run through safe_initiative to enforce defaults after edits.
        updated = safe_initiative(updated)

        try:
            supabase.table("initiatives").update(updated).eq("id", init_id).execute()
        except Exception as e:
            return jsonify({"error": f"Could not update initiative: {e}"}), 500

        with cache_lock:
            cache["initiatives"] = [updated if i["id"] == init_id else i
                                    for i in cache["initiatives"]]

        changed = [f for f in updatable if f in data]
        log_activity(data.get("user", "System"), "Initiative Edited",
                     init_id, updated.get("name", ""),
                     f"Changed: {', '.join(changed)}")
        return jsonify({"message": "Initiative updated.",
                        "initiative": updated}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/initiatives/<init_id>", methods=["DELETE"])
def delete_initiative(init_id):
    try:
        data = request.get_json(silent=True) or {}
        # Fallback to query string so callers can DELETE without a body.
        req_user = (data.get("user") or request.args.get("user") or "System").strip()

        with cache_lock:
            init = next((i for i in cache.get("initiatives", []) if i["id"] == init_id), None)
        if not init:
            return jsonify({"error": "Initiative not found."}), 404

        # 1. Clear initiative_id off every task that points at this id.
        try:
            supabase.table("tasks").update({"initiative_id": ""}).eq("initiative_id", init_id).execute()
        except Exception as e:
            # Don't abort — we'd rather end up with orphaned tasks
            # pointing at a deleted initiative than block the delete.
            print(f"[delete_initiative] task unlink warning: {e}")

        # 2. Delete the initiative row.
        try:
            supabase.table("initiatives").delete().eq("id", init_id).execute()
        except Exception as e:
            return jsonify({"error": f"Could not delete initiative: {e}"}), 500

        # 3. Reflect both changes in the in-memory cache.
        unlinked = 0
        with cache_lock:
            cache["initiatives"] = [i for i in cache["initiatives"] if i["id"] != init_id]
            for t in cache.get("tasks", []):
                if (t.get("initiative_id") or "") == init_id:
                    t["initiative_id"] = ""
                    unlinked += 1

        log_activity(req_user, "Initiative Deleted", init_id,
                     init.get("name", ""),
                     f"Unlinked {unlinked} task(s)")
        return jsonify({"message": "Initiative deleted.",
                        "unlinked_tasks": unlinked}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: ACTIVITY LOG ────────────────────────────────────────────
@app.route("/api/activity", methods=["GET"])
def get_activity():
    try:
        limit = int(request.args.get("limit", 50))
        task_id = request.args.get("task_id", "").strip()
        if task_id:
            # Query Supabase directly — the in-memory cache is capped at 200
            # entries across ALL tasks, so old creation events fall off.
            res = supabase.table("activity_log").select("*") \
                .eq("task_id", task_id) \
                .order("id", desc=True).limit(limit).execute()
            activity = [
                {
                    "timestamp":  a.get("timestamp", ""),
                    "username":   a.get("username", ""),
                    "action":     a.get("action", ""),
                    "task_id":    a.get("task_id", ""),
                    "task_title": a.get("task_title", ""),
                    "details":    a.get("details", ""),
                }
                for a in (res.data or [])
            ]
            return jsonify({"activity": activity}), 200
        with cache_lock:
            activity = list(cache.get("activity", []))
        return jsonify({"activity": activity[:limit]}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: AUTO-ASSIGN ─────────────────────────────────────────────
@app.route("/api/auto-assign", methods=["POST"])
def auto_assign():
    try:
        data = request.get_json()
        titles = data.get("titles", [])
        if not titles:
            return jsonify({"error": "No titles provided."}), 400
        if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
            return jsonify({"error": "Groq API key not configured."}), 500

        with cache_lock:
            all_tasks = list(cache["tasks"])
        examples = [{"title": t["title"], "function": t["team"], "sub_category": t["sub_category"]} for t in all_tasks if t["team"] and t["sub_category"]][:50]

        prompt = f"""You are a task classifier for LingoTran, a language learning platform.
AVAILABLE FUNCTIONS AND SUB-CATEGORIES:
{json.dumps(FUNCTIONS, indent=2)}
EXAMPLES FROM EXISTING TASKS:
{json.dumps(examples, indent=2)}
Classify each new task title into the best matching function and sub_category.
Respond ONLY with a JSON array. Each item must have: title, function, sub_category
New task titles: {json.dumps(titles)}"""

        response = http_requests.post("https://api.groq.com/openai/v1/chat/completions",
            json={"model":"llama-3.1-8b-instant","messages":[{"role":"system","content":"JSON-only classifier."},{"role":"user","content":prompt}],"temperature":0.2,"max_tokens":2000},
            headers={"Authorization":f"Bearer {GROQ_API_KEY}","Content-Type":"application/json"}, timeout=30)
        answer = response.json()["choices"][0]["message"]["content"].strip()
        if answer.startswith("```"):
            answer = answer.split("\n",1)[-1].rsplit("```",1)[0].strip()
        assignments = json.loads(answer)
        # Same default-to-creator rule as add_task / bulk_add_tasks: if
        # the model didn't choose an assignee (or said "Unassigned"),
        # assign to whoever invoked auto-assign.
        req_user = (data.get("user") or "").strip()
        for a in assignments:
            if a.get("function") not in FUNCTIONS: a["function"] = ""
            if a.get("sub_category") not in FUNCTIONS.get(a.get("function",[]),[]):
                subs = FUNCTIONS.get(a.get("function"),[])
                a["sub_category"] = subs[0] if subs else ""
            cur = (a.get("assigned_to") or "").strip()
            if not cur or cur.lower() == "unassigned":
                a["assigned_to"] = req_user
        return jsonify({"assignments": assignments}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: AI INSIGHTS ─────────────────────────────────────────────
@app.route("/api/insights", methods=["GET"])
def get_insights():
    try:
        if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
            return jsonify({"insights": []}), 200
        with cache_lock:
            all_tasks = list(cache["tasks"])
        today = datetime.now().strftime("%Y-%m-%d")
        open_tasks = [t for t in all_tasks if t["status"] != "done"]
        if not open_tasks:
            return jsonify({"insights": []}), 200
        task_lines = [f"Title:{t['title']} | Owner:{t['assigned_to'] or 'Unassigned'} | Function:{t['team']} | Priority:{t['priority']} | Status:{t['status']} | Due:{t['due_date'] or 'No date'} {'OVERDUE' if t['due_date'] and t['due_date'] < today else ''}" for t in open_tasks]
        total=len(all_tasks); done=sum(1 for t in all_tasks if t["status"]=="done"); overdue=sum(1 for t in all_tasks if t["status"]!="done" and t["due_date"] and t["due_date"]<today); blocked=sum(1 for t in all_tasks if t["status"] in ["blocked","hold"])
        workload = {}
        for t in all_tasks:
            if t["status"]!="done" and t["assigned_to"] and t["assigned_to"]!="Unassigned":
                workload[t["assigned_to"]]=workload.get(t["assigned_to"],0)+1
        prompt = f"""You are a smart task management assistant for LingoTran.
Today: {today}. Stats: Total:{total}|Done:{done}|Overdue:{overdue}|Blocked:{blocked}. Workload: {json.dumps(workload)}
OPEN TASKS:\n{chr(10).join(task_lines[:50])}
Generate exactly 3 short actionable notifications (1 sentence each, max 15 words, different types, add emoji).
Respond ONLY with a JSON array of 3 strings."""
        response = http_requests.post("https://api.groq.com/openai/v1/chat/completions",
            json={"model":"llama-3.1-8b-instant","messages":[{"role":"system","content":"Return only valid JSON arrays of strings."},{"role":"user","content":prompt}],"temperature":0.4,"max_tokens":300},
            headers={"Authorization":f"Bearer {GROQ_API_KEY}","Content-Type":"application/json"}, timeout=15)
        answer = response.json()["choices"][0]["message"]["content"].strip()
        if answer.startswith("```"): answer = answer.split("\n",1)[-1].rsplit("```",1)[0].strip()
        insights = [str(i) for i in json.loads(answer)[:3]]
        return jsonify({"insights": insights}), 200
    except Exception as e:
        print(f"Insights error: {e}")
        return jsonify({"insights": []}), 200

# ── API: ASK AI ──────────────────────────────────────────────────
@app.route("/api/ask", methods=["POST"])
def ask_ai():
    try:
        data = request.get_json()
        question = data.get("question", "").strip()
        if not question:
            return jsonify({"error": "Question is required."}), 400
        if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
            return jsonify({"error": "No AI API key configured."}), 500

        with cache_lock:
            all_tasks = list(cache["tasks"])
            all_users = list(cache["users"])
            all_teams = list(cache["teams"])

        today = datetime.now().strftime("%Y-%m-%d")
        task_lines = [f"ID:{t['task_id']}|Title:{t['title']}|Owner:{t['assigned_to']}|Function:{t['team']}|Priority:{t['priority']}|Status:{t['status']}|Due:{t['due_date'] or 'No date'}{'|OVERDUE' if t['due_date'] and t['status']!='done' and t['due_date']<today else ''}" for t in all_tasks]
        total=len(all_tasks); done=sum(1 for t in all_tasks if t["status"]=="done"); in_progress=sum(1 for t in all_tasks if t["status"] in ["progress","just"]); blocked=sum(1 for t in all_tasks if t["status"] in ["blocked","hold"]); overdue_count=sum(1 for t in all_tasks if t["status"]!="done" and t["due_date"] and t["due_date"]<today)

        system_prompt = f"""You are TaskFlow AI Assistant. Today: {today}.
TEAM MEMBERS: {', '.join(u['name']+'('+u['team']+')' for u in all_users)}
STATS: Total:{total}|Done:{done}|In Progress:{in_progress}|Blocked:{blocked}|Overdue:{overdue_count}
ALL TASKS:\n{chr(10).join(task_lines)}
INSTRUCTIONS: Only answer questions about tasks, team workload, deadlines, and project management. 
If asked anything unrelated say "I can only help with task-related questions."
Be concise and data-driven."""

        response = http_requests.post("https://api.groq.com/openai/v1/chat/completions",
            json={"model":"llama-3.1-8b-instant","messages":[{"role":"system","content":system_prompt},{"role":"user","content":question}],"temperature":0.3,"max_tokens":1000},
            headers={"Authorization":f"Bearer {GROQ_API_KEY}","Content-Type":"application/json"}, timeout=30)
        answer = response.json()["choices"][0].get("message",{}).get("content","")
        return jsonify({"answer": answer or "Sorry, I could not generate a response."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: REFRESH ─────────────────────────────────────────────────
@app.route("/api/refresh", methods=["POST"])
def refresh_cache():
    sync_cache_from_supabase()
    return jsonify({"message": "Cache refreshed.", "last_sync": cache["last_sync"]}), 200

# ── STARTUP ──────────────────────────────────────────────────────
def startup():
    print("Seeding default teams...")
    seed_teams()
    print("Loading data into cache...")
    sync_cache_from_supabase()
    sync_thread = threading.Thread(target=background_sync, daemon=True)
    sync_thread.start()
    print("Background sync started.")

startup()

if __name__ == "__main__":
    print("Starting TaskFlow on http://127.0.0.1:5000")
    app.run(debug=True, host="0.0.0.0", port=5000, use_reloader=False)