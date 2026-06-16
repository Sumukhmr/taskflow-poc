from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from supabase import create_client, Client
from dotenv import load_dotenv
from datetime import datetime
import uuid
import threading
import time
import os
import io
import csv
import json
import requests as http_requests

load_dotenv()

app = Flask(__name__)
CORS(app)

# ── CONFIGURATION ───────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

# ── SUPABASE CLIENT ─────────────────────────────────────────────
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── CONSTANTS ───────────────────────────────────────────────────
# Built-in defaults — seed values only. FUNCTIONS is rebuilt at runtime
# from Supabase (teams + sub_categories tables) so the user can add their
# own functions and sub-categories without code changes.
DEFAULT_FUNCTIONS = {
    "Sales":       ["Enterprise Pipeline", "Outbound", "Inbound", "Account Management", "Partnerships"],
    "Content":     ["Pedagogy", "Content Ingestion", "Content Review/Testing", "Curriculum Design"],
    "Engineering": ["Backend", "Frontend", "Infrastructure/DevOps", "Mobile", "Data"],
    "Operations":  ["Customer Onboarding", "Process", "Vendor Management", "Internal Tools"],
    "Marketing":   ["Campaigns", "Events/Conferences", "Content Marketing", "SEO", "Social", "Brand"],
    "Finance":     ["Accounting", "Budgeting", "Payroll", "Reporting"],
    "HR":          ["Recruiting", "People Ops", "Learning & Development", "Culture"],
    "Product":     ["Feature", "Enhancement", "UI/UX", "Bug/Issue", "Analytics/Reporting", "Research"],
    "Legal":       ["Contracts", "Compliance", "IP"],
    "Support":     ["Customer Support", "Technical Support"],
}
# FUNCTIONS is mutated in place by sync_functions() so existing references
# (e.g. FUNCTIONS["Engineering"]) keep working.
FUNCTIONS = {k: list(v) for k, v in DEFAULT_FUNCTIONS.items()}
DEFAULT_TEAMS = list(DEFAULT_FUNCTIONS.keys())
VALID_STATUSES = [
    "todo", "just", "progress", "hold", "blocked", "review",
    "pushed_dev", "pushed_uat", "pushed_cloud",
    "tested_dev", "tested_uat", "tested_cloud",
    "done",
]
VALID_PRIORITIES = ["high", "medium", "low"]
VALID_QUADRANTS = ["q1", "q2", "q3", "q4"]

# ── IN-MEMORY CACHE ─────────────────────────────────────────────
cache = {
    "tasks": [],
    "teams": [],
    "users": [],
    "activity": [],
    "functions": FUNCTIONS,
    "last_sync": None
}
cache_lock = threading.Lock()

# ── HELPERS ──────────────────────────────────────────────────────
def generate_task_id():
    return f"TSK-{uuid.uuid4().hex[:6].upper()}"

def get_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_date(value):
    """Convert common user-typed / spreadsheet date formats to YYYY-MM-DD.

    Handles M/D/YYYY, MM-DD-YYYY, DD/MM/YYYY, DD-MM-YYYY, YYYY/M/D,
    M/D/YY, DD/MM/YY, and the canonical YYYY-MM-DD. Returns "" for empty
    input. If parsing fails for every format, returns the original string
    so the caller can decide what to do (we don't silently drop the
    user's data)."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    # Already ISO? Trust it.
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return s
    except ValueError:
        pass

    # Try common formats in order. NOTE: M/D/YYYY and DD/MM/YYYY are
    # ambiguous when day <= 12, so we try US-style first (more common
    # in spreadsheets exported from Excel/Google Sheets in en-US locale),
    # then European-style as a fallback.
    formats = [
        "%m/%d/%Y", "%m-%d-%Y",
        "%d/%m/%Y", "%d-%m-%Y",
        "%Y/%m/%d",
        "%m/%d/%y", "%d/%m/%y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return s  # fall back to original — unparseable but preserved

def write_in_background(fn, *args, **kwargs):
    def wrapper():
        try:
            fn(*args, **kwargs)
        except Exception as e:
            print(f"Background write failed: {e}")
    thread = threading.Thread(target=wrapper, daemon=True)
    thread.start()

def log_activity(user, action, task_id="", task_title="", details=""):
    row = {
        "timestamp": get_timestamp(),
        "username": user or "System",
        "action": action,
        "task_id": task_id,
        "task_title": task_title,
        "details": details
    }
    with cache_lock:
        cache.setdefault("activity", []).insert(0, row)
        cache["activity"] = cache["activity"][:200]

    def sheet_write():
        supabase.table("activity_log").insert(row).execute()
    write_in_background(sheet_write)

def safe_task(t):
    """Ensure task dict has all fields with proper defaults."""
    return {
        "task_id":      t.get("task_id", ""),
        "title":        t.get("title", ""),
        "description":  t.get("description", "") or "",
        "assigned_to":  t.get("assigned_to", "") or "",
        "team":         t.get("team", "") or "",
        "due_date":     t.get("due_date", "") or "",
        "priority":     t.get("priority", "medium") or "medium",
        "status":       t.get("status", "todo") or "todo",
        "created_by":   t.get("created_by", "") or "",
        "created_at":   t.get("created_at", "") or "",
        "updated_at":   t.get("updated_at", "") or "",
        "sub_category": t.get("sub_category", "") or "",
        "quadrant":     t.get("quadrant", "q2") or "q2",
        "notes":        t.get("notes", "") or "",
        "type":         t.get("type", "") or "",
        "start_date":   t.get("start_date", "") or "",
        "attachments":  t.get("attachments") or [],
    }

# ── CACHE SYNC ──────────────────────────────────────────────────
def sync_cache_from_supabase():
    try:
        # Load tasks
        res = supabase.table("tasks").select("*").execute()
        tasks = [safe_task(t) for t in (res.data or [])]

        # Load teams
        res = supabase.table("teams").select("*").execute()
        teams = [r["team_name"] for r in (res.data or [])]
        if not teams:
            teams = DEFAULT_TEAMS

        # Load users
        res = supabase.table("users").select("*").execute()
        users = []
        for u in (res.data or []):
            users.append({
                "name":     u.get("name", ""),
                "team":     u.get("team", "") or "",
                "email":    u.get("email", "") or "",
                "password": u.get("password", "") or ""
            })

        # Load activity
        res = supabase.table("activity_log").select("*").order("id", desc=True).limit(200).execute()
        activity = []
        for a in (res.data or []):
            activity.append({
                "timestamp":  a.get("timestamp", ""),
                "user":       a.get("username", ""),
                "action":     a.get("action", ""),
                "task_id":    a.get("task_id", ""),
                "task_title": a.get("task_title", ""),
                "details":    a.get("details", "")
            })

        with cache_lock:
            cache["tasks"] = tasks
            cache["teams"] = teams
            cache["users"] = users
            cache["activity"] = activity
            cache["last_sync"] = get_timestamp()

        # Rebuild FUNCTIONS so user-added functions / sub-categories show up.
        sync_functions(teams_hint=teams)

        print(f"Cache synced: {len(tasks)} tasks, {len(teams)} teams, {len(users)} users, {len(FUNCTIONS)} functions")

    except Exception as e:
        print(f"Cache sync failed: {e}")


# ── FUNCTIONS / SUB-CATEGORIES (user-extendable) ────────────────
def _load_sub_categories():
    """Read the sub_categories overlay from Supabase. Returns a
    {function: [sub, ...]} dict, or None if the table isn't there yet."""
    try:
        res = supabase.table("sub_categories").select("*").execute()
        out = {}
        for row in (res.data or []):
            fn = (row.get("function") or "").strip()
            name = (row.get("sub_category") or row.get("name") or "").strip()
            if fn and name:
                out.setdefault(fn, []).append(name)
        return out
    except Exception as e:
        print(f"sub_categories table not available (using built-in defaults only): {e}")
        return None


def sync_functions(teams_hint=None):
    """Rebuild the FUNCTIONS dict from teams + sub_categories tables.
    Built-in defaults are still used for any team that has no overlay row,
    so the existing seeded functions keep their sub-categories without
    requiring rows in the sub_categories table."""
    try:
        if teams_hint is None:
            tres = supabase.table("teams").select("team_name").execute()
            teams = [r["team_name"] for r in (tres.data or [])]
        else:
            teams = list(teams_hint)
        if not teams:
            teams = list(DEFAULT_FUNCTIONS.keys())

        new_fn = {}
        for fn in teams:
            # Start from built-in defaults if we know this function, else empty
            new_fn[fn] = list(DEFAULT_FUNCTIONS.get(fn, []))

        # Overlay any user-added sub-categories
        overlay = _load_sub_categories()
        if overlay:
            for fn, subs in overlay.items():
                bucket = new_fn.setdefault(fn, [])
                for s in subs:
                    if s not in bucket:
                        bucket.append(s)

        FUNCTIONS.clear()
        FUNCTIONS.update(new_fn)
        with cache_lock:
            cache["functions"] = FUNCTIONS
    except Exception as e:
        print(f"sync_functions failed (keeping previous): {e}")

def background_sync():
    while True:
        time.sleep(60)
        sync_cache_from_supabase()

# ── SEED DEFAULT TEAMS ──────────────────────────────────────────
def seed_teams():
    try:
        res = supabase.table("teams").select("team_name").execute()
        existing = [r["team_name"] for r in (res.data or [])]
        new_teams = [{"team_name": t} for t in DEFAULT_TEAMS if t not in existing]
        if new_teams:
            supabase.table("teams").insert(new_teams).execute()
            print(f"Seeded teams: {[t['team_name'] for t in new_teams]}")
    except Exception as e:
        print(f"Team seeding failed: {e}")

# ── SERVE FRONTEND ──────────────────────────────────────────────
@app.route("/")
def serve_frontend():
    return send_from_directory(".", "index.html")

# ── API: TEAMS ──────────────────────────────────────────────────
@app.route("/api/teams", methods=["GET"])
def get_teams():
    with cache_lock:
        teams = list(cache["teams"])
    return jsonify({"teams": teams}), 200

# ── API: USERS ──────────────────────────────────────────────────
@app.route("/api/users", methods=["GET"])
def get_users():
    with cache_lock:
        users = list(cache["users"])
    safe = [{"name": u["name"], "team": u["team"], "email": u.get("email", "")} for u in users]
    return jsonify({"users": safe}), 200

# ── API: LOGIN ──────────────────────────────────────────────────
def get_assigned_open_tasks(username):
    priority_order = {"high": 0, "medium": 1, "low": 2}
    uname_lower = username.lower()
    with cache_lock:
        result = [
            {k: t[k] for k in ["task_id","title","priority","status","due_date","created_by","created_at","team","sub_category"]}
            for t in cache["tasks"]
            if t["assigned_to"].lower() == uname_lower and t["status"] != "done"
        ]
    result.sort(key=lambda t: (priority_order.get(t["priority"], 3), t["due_date"] or "9999"))
    return result

@app.route("/api/login", methods=["POST"])
def login():
    try:
        data = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        if not username or not password:
            return jsonify({"error": "Username and password required."}), 400
        with cache_lock:
            user = next((u for u in cache["users"] if u["name"].lower() == username.lower() and u["password"] == password), None)
        if not user:
            return jsonify({"error": "Invalid username or password."}), 401
        log_activity(user["name"], "Logged In", "", "", "")
        return jsonify({
            "message": "Login successful.",
            "user": {"name": user["name"], "team": user["team"]},
            "assigned_tasks": get_assigned_open_tasks(user["name"])
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: SIGNUP ──────────────────────────────────────────────────
@app.route("/api/signup", methods=["POST"])
def signup():
    try:
        data = request.get_json()
        name     = data.get("name", "").strip()
        email    = data.get("email", "").strip()
        password = data.get("password", "").strip()
        if not name or not password:
            return jsonify({"error": "Name and password are required."}), 400
        if len(password) < 4:
            return jsonify({"error": "Password must be at least 4 characters."}), 400
        with cache_lock:
            existing = next((u for u in cache["users"] if u["name"].lower() == name.lower()), None)
        if existing:
            return jsonify({"error": "Username already taken. Please log in instead."}), 409
        new_user = {"name": name, "team": "", "email": email, "password": password}
        with cache_lock:
            cache["users"].append(new_user)
        def sheet_write():
            supabase.table("users").insert({"name": name, "team": "", "email": email, "password": password}).execute()
        write_in_background(sheet_write)
        log_activity(name, "Signed Up", "", "", f"Email: {email}")
        return jsonify({"message": "Account created successfully.", "user": {"name": name, "team": ""}}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: MY TASKS ────────────────────────────────────────────────
@app.route("/api/my-tasks", methods=["GET"])
def my_tasks():
    try:
        username = request.args.get("user", "").strip()
        if not username:
            return jsonify({"tasks": []}), 200
        return jsonify({"tasks": get_assigned_open_tasks(username)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: FUNCTIONS ───────────────────────────────────────────────
@app.route("/api/functions", methods=["GET"])
def get_functions():
    return jsonify({"functions": FUNCTIONS}), 200


@app.route("/api/functions/add", methods=["POST"])
def add_function():
    """Add a new top-level function (stored as a team)."""
    try:
        data = request.get_json(force=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Function name is required."}), 400
        if len(name) > 60:
            return jsonify({"error": "Function name is too long (max 60)."}), 400

        # Case-insensitive duplicate check against the current cache
        existing = {t.lower() for t in cache.get("teams", [])}
        if name.lower() in existing:
            return jsonify({"error": f"'{name}' already exists."}), 409

        try:
            supabase.table("teams").insert({"team_name": name}).execute()
        except Exception as e:
            return jsonify({"error": f"Could not save function: {e}"}), 500

        with cache_lock:
            if name not in cache["teams"]:
                cache["teams"].append(name)
        sync_functions(teams_hint=cache["teams"])

        log_activity(data.get("user", "System"), "Function Added", "", "", f"Function: {name}")
        return jsonify({"success": True, "function": name, "functions": FUNCTIONS}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/sub-categories/add", methods=["POST"])
def add_sub_category():
    """Add a new sub-category under an existing function."""
    try:
        data = request.get_json(force=True) or {}
        function = (data.get("function") or "").strip()
        sub_category = (data.get("sub_category") or data.get("name") or "").strip()

        if not function:
            return jsonify({"error": "Function is required."}), 400
        if not sub_category:
            return jsonify({"error": "Sub-category name is required."}), 400
        if len(sub_category) > 80:
            return jsonify({"error": "Sub-category name is too long (max 80)."}), 400
        if function not in FUNCTIONS and function not in cache.get("teams", []):
            return jsonify({"error": f"Unknown function '{function}'. Add the function first."}), 400

        # Duplicate guard (case-insensitive within the function)
        existing = {s.lower() for s in FUNCTIONS.get(function, [])}
        if sub_category.lower() in existing:
            return jsonify({"error": f"'{sub_category}' already exists under {function}."}), 409

        try:
            supabase.table("sub_categories").insert({
                "function": function,
                "sub_category": sub_category,
            }).execute()
        except Exception as e:
            msg = str(e)
            hint = ""
            lower = msg.lower()
            if "sub_categories" in lower and (
                "not find" in lower or "relation" in lower or "does not exist" in lower or "schema cache" in lower
            ):
                hint = (
                    " — Create this table in Supabase first (SQL editor): "
                    "CREATE TABLE sub_categories (id bigserial PRIMARY KEY, "
                    "function text NOT NULL, sub_category text NOT NULL, "
                    "created_at timestamptz DEFAULT now(), "
                    "UNIQUE(function, sub_category));"
                )
            return jsonify({"error": f"Could not save sub-category: {msg}{hint}"}), 500

        # Update FUNCTIONS in place so the response reflects the new state
        FUNCTIONS.setdefault(function, []).append(sub_category)
        with cache_lock:
            cache["functions"] = FUNCTIONS

        log_activity(data.get("user", "System"), "Sub-category Added", "", "",
                     f"{function} / {sub_category}")
        return jsonify({
            "success": True,
            "function": function,
            "sub_category": sub_category,
            "functions": FUNCTIONS,
        }), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: STATS ───────────────────────────────────────────────────
@app.route("/api/stats", methods=["GET"])
def get_stats():
    try:
        with cache_lock:
            all_tasks = list(cache["tasks"])
            teams = list(cache["teams"])

        today = datetime.now().strftime("%Y-%m-%d")
        filtered = all_tasks
        total = len(filtered)
        done = sum(1 for t in filtered if t["status"] == "done")
        in_progress = sum(1 for t in filtered if t["status"] in ["progress", "just"])
        blocked = sum(1 for t in filtered if t["status"] in ["blocked", "hold"])
        unassigned = sum(1 for t in filtered if not t["assigned_to"] or t["assigned_to"] == "Unassigned")
        overdue = sum(1 for t in filtered if t["status"] != "done" and t["due_date"] and t["due_date"] < today)
        complete_pct = round((done / total * 100) if total > 0 else 0)

        by_function = []
        for team in teams:
            tt = [t for t in filtered if t["team"] == team]
            if tt:
                by_function.append({"team": team, "total": len(tt), "completed": sum(1 for t in tt if t["status"] == "done")})

        by_quadrant = {q: sum(1 for t in filtered if t["quadrant"] == q and t["status"] != "done") for q in VALID_QUADRANTS}
        by_status = [{"id": s, "count": sum(1 for t in filtered if t["status"] == s)} for s in VALID_STATUSES if sum(1 for t in filtered if t["status"] == s) > 0]

        return jsonify({
            "total": total, "done": done, "in_progress": in_progress,
            "blocked": blocked, "overdue": overdue, "unassigned": unassigned,
            "complete_pct": complete_pct, "by_function": by_function,
            "by_quadrant": by_quadrant, "by_status": by_status
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: GET TASKS ───────────────────────────────────────────────
@app.route("/api/tasks", methods=["GET"])
def get_tasks():
    try:
        with cache_lock:
            tasks = list(cache["tasks"])
        team = request.args.get("team", "")
        sub = request.args.get("sub", "")
        status = request.args.get("status", "")
        priority = request.args.get("priority", "")
        quadrant = request.args.get("quadrant", "")
        owner = request.args.get("owner", "")
        if team and team != "All": tasks = [t for t in tasks if t["team"] == team]
        if sub and sub != "all": tasks = [t for t in tasks if t["sub_category"] == sub]
        if status and status != "all": tasks = [t for t in tasks if t["status"] == status]
        if priority and priority != "all": tasks = [t for t in tasks if t["priority"] == priority]
        if quadrant and quadrant != "all": tasks = [t for t in tasks if t["quadrant"] == quadrant]
        if owner and owner != "all": tasks = [t for t in tasks if t["assigned_to"] == owner]
        return jsonify({"tasks": tasks}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: ADD SINGLE TASK ─────────────────────────────────────────
@app.route("/api/tasks", methods=["POST"])
def add_task():
    try:
        data = request.get_json()
        if not data.get("title", "").strip():
            return jsonify({"error": "'title' is required."}), 400

        priority = data.get("priority", "medium")
        if priority not in VALID_PRIORITIES: priority = "medium"
        status = data.get("status", "todo")
        if status not in VALID_STATUSES: status = "todo"
        quadrant = data.get("quadrant", "q2")
        if quadrant not in VALID_QUADRANTS: quadrant = "q2"
        team = data.get("team", "").strip()
        sub_category = data.get("sub_category", "").strip()
        if team and team in FUNCTIONS and not sub_category:
            _subs = FUNCTIONS.get(team) or []
            if _subs:
                sub_category = _subs[0]

        task_id = generate_task_id()
        now = get_timestamp()
        new_task = safe_task({
            "task_id": task_id,
            "title": data["title"].strip(),
            "description": data.get("description", "").strip(),
            "assigned_to": data.get("assigned_to", "").strip(),
            "team": team,
            "due_date": normalize_date(data.get("due_date", "")),
            "priority": priority,
            "status": status,
            "created_by": data.get("user", "System"),
            "created_at": now,
            "updated_at": now,
            "sub_category": sub_category,
            "quadrant": quadrant,
            "notes": data.get("notes", "").strip(),
            "type": data.get("type", "").strip(),
            "start_date": normalize_date(data.get("start_date", "")),
            "attachments": data.get("attachments", []),
        })

        with cache_lock:
            cache["tasks"].append(new_task)

        def db_write():
            supabase.table("tasks").insert(new_task).execute()
        write_in_background(db_write)
        log_activity(data.get("user", "System"), "Created", task_id, new_task["title"], f"Function: {team}, Priority: {priority}")
        return jsonify({"message": "Task created successfully.", "task": new_task}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: BULK ADD TASKS ──────────────────────────────────────────
@app.route("/api/tasks/bulk", methods=["POST"])
def bulk_add_tasks():
    try:
        data = request.get_json()
        titles = [t.strip() for t in data.get("titles", []) if t.strip()]
        if not titles:
            return jsonify({"error": "No task titles provided."}), 400

        priority = data.get("priority", "medium")
        if priority not in VALID_PRIORITIES: priority = "medium"
        status = data.get("status", "todo")
        if status not in VALID_STATUSES: status = "todo"
        quadrant = data.get("quadrant", "q2")
        if quadrant not in VALID_QUADRANTS: quadrant = "q2"
        team = data.get("team", "").strip()
        sub_category = data.get("sub_category", "").strip()
        if team and team in FUNCTIONS and not sub_category:
            _subs = FUNCTIONS.get(team) or []
            if _subs:
                sub_category = _subs[0]
        assigned_to = data.get("assigned_to", "").strip()
        due_date = normalize_date(data.get("due_date", ""))
        start_date = normalize_date(data.get("start_date", ""))
        type_val = data.get("type", "").strip()
        now = get_timestamp()

        new_tasks = []
        for title in titles:
            task_id = generate_task_id()
            new_tasks.append(safe_task({
                "task_id": task_id, "title": title, "description": "",
                "assigned_to": assigned_to, "team": team, "due_date": due_date,
                "priority": priority, "status": status,
                "created_by": data.get("user", "System"),
                "created_at": now, "updated_at": now,
                "sub_category": sub_category, "quadrant": quadrant,
                "notes": "", "type": type_val, "start_date": start_date, "attachments": []
            }))

        with cache_lock:
            cache["tasks"].extend(new_tasks)

        def db_write():
            supabase.table("tasks").insert(new_tasks).execute()
        write_in_background(db_write)
        log_activity(data.get("user", "System"), "Bulk Created", "", f"{len(new_tasks)} tasks", f"Function: {team}")
        return jsonify({"message": f"{len(new_tasks)} tasks created.", "tasks": new_tasks, "count": len(new_tasks)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: IMPORT TASKS FROM CSV OR XLSX ──────────────────────────
def _normalize_header(h):
    """Lowercase, trim, drop parenthetical hints like '(YYYY-MM-DD)',
    and replace spaces/dashes with underscores so the template header
    'due_date (YYYY-MM-DD)' still maps to 'due_date'."""
    s = (h or "").strip()
    if "(" in s:
        s = s.split("(", 1)[0]
    return s.strip().lower().replace(" ", "_").replace("-", "_")


def _coerce_cell(v):
    """Coerce a cell value (CSV string, XLSX number, XLSX date, etc.) to
    a plain string. Real date/datetime objects become ISO YYYY-MM-DD."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):           # datetime.date / datetime.datetime
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_csv_rows(raw):
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(content))
    return list(reader.fieldnames or []), [dict(r) for r in reader]


def _parse_xlsx_rows(raw):
    try:
        import openpyxl  # noqa - imported lazily so .csv-only deploys keep working
    except ImportError:
        raise RuntimeError("openpyxl is required to import .xlsx files. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in all_rows[0]]
    out = []
    for row in all_rows[1:]:
        if all(c in (None, "") for c in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if h:
                d[h] = _coerce_cell(v)
        out.append(d)
    return headers, out


@app.route("/api/tasks/import-csv", methods=["POST"])
def import_tasks_csv():
    """Bulk-create tasks from an uploaded .csv OR .xlsx file.
    Required column: title. All other columns are optional and fall
    back to the same defaults the regular /api/tasks endpoint uses.
    Date columns may include format hints like 'due_date (YYYY-MM-DD)'."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded (expected multipart field 'file')."}), 400
        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"error": "Empty filename."}), 400

        fname = f.filename.lower()
        raw = f.read()
        try:
            if fname.endswith(".xlsx"):
                fieldnames, rows = _parse_xlsx_rows(raw)
            elif fname.endswith(".csv"):
                fieldnames, rows = _parse_csv_rows(raw)
            else:
                return jsonify({"error": "Only .csv and .xlsx files are accepted."}), 400
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 500
        except Exception as e:
            return jsonify({"error": f"Could not parse file: {e}"}), 400

        if not fieldnames:
            return jsonify({"error": "File is empty or has no header row."}), 400

        # Case-insensitive header mapping (also strips format hints in parens).
        field_map = {}
        for fn in fieldnames:
            key = _normalize_header(fn)
            if key:
                field_map[key] = fn
        if "title" not in field_map:
            return jsonify({
                "error": "File must have a 'title' column.",
                "found_columns": list(fieldnames),
            }), 400

        def cell(row, name):
            src = field_map.get(name)
            if not src:
                return ""
            v = row.get(src, "")
            return (v or "").strip() if isinstance(v, str) else _coerce_cell(v)

        now = get_timestamp()
        user = (request.form.get("user") or "System").strip() or "System"

        created = []
        skipped_no_title = 0
        errors = []

        for i, row in enumerate(rows, start=2):  # row 2 = first data row
            title = cell(row, "title")
            if not title:
                skipped_no_title += 1
                continue
            try:
                priority = (cell(row, "priority") or "medium").lower()
                if priority not in VALID_PRIORITIES:
                    priority = "medium"
                status = (cell(row, "status") or "todo").lower()
                if status not in VALID_STATUSES:
                    status = "todo"
                quadrant = (cell(row, "quadrant") or "q2").lower()
                if quadrant not in VALID_QUADRANTS:
                    quadrant = "q2"

                team = cell(row, "team")
                sub_category = cell(row, "sub_category")
                if team and team in FUNCTIONS and not sub_category:
                    _subs = FUNCTIONS.get(team) or []
                    if _subs:
                        sub_category = _subs[0]

                new_task = safe_task({
                    "task_id":      generate_task_id(),
                    "title":        title,
                    "description":  cell(row, "description"),
                    "assigned_to":  cell(row, "assigned_to"),
                    "team":         team,
                    "due_date":     normalize_date(cell(row, "due_date")),
                    "priority":     priority,
                    "status":       status,
                    "created_by":   cell(row, "created_by") or user,
                    "created_at":   now,
                    "updated_at":   now,
                    "sub_category": sub_category,
                    "quadrant":     quadrant,
                    "notes":        cell(row, "notes"),
                    "type":         cell(row, "type"),
                    "start_date":   normalize_date(cell(row, "start_date")),
                    "attachments":  [],
                })
                created.append(new_task)
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        if created:
            with cache_lock:
                cache["tasks"].extend(created)

            def db_write(rows=created):
                supabase.table("tasks").insert(rows).execute()
            write_in_background(db_write)

            log_activity(
                user, "CSV Import", "", f"{len(created)} tasks",
                f"from {f.filename}"
            )

        return jsonify({
            "success": True,
            "created": len(created),
            "skipped": skipped_no_title,
            "errors": errors[:10],
            "filename": f.filename,
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: DOWNLOAD XLSX TEMPLATE ─────────────────────────────────
@app.route("/api/tasks/template.xlsx", methods=["GET"])
def download_xlsx_template():
    """Generate and serve an .xlsx template with:
    - Date-formatted cells on due_date / start_date (Excel shows the
      native calendar picker when the user double-clicks the cell)
    - Drop-down lists on status / priority / quadrant
    - Frozen header row + bold purple header
    - One example data row so the user sees the shape immediately"""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"error": "openpyxl is not installed on this server. Run: pip install openpyxl"}), 500
        from datetime import date

        headers = [
            "title", "description", "team", "sub_category", "assigned_to",
            "priority", "status", "quadrant",
            "due_date", "start_date", "type", "notes",
        ]
        DATE_COLS = {"due_date", "start_date"}
        col_index = {h: i + 1 for i, h in enumerate(headers)}

        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill("solid", fgColor="5B21B6")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell_ in ws[1]:
            cell_.fill = header_fill
            cell_.font = header_font
            cell_.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Example row — gives the user an immediate visual cue for each column
        ws.append([
            "Example task — only title is required",
            "Optional description",
            "Development",
            "Backend",
            "Alice, Bob",
            "medium",
            "todo",
            "q2",
            date(2026, 6, 30),
            date(2026, 6, 15),
            "task",
            "Any notes here",
        ])

        # Set column widths + apply date format down each date column
        for h, col in col_index.items():
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = max(len(h) + 4, 18)
            if h in DATE_COLS:
                # Format up to row 1000 so users see the picker as they fill it in
                for r in range(2, 1001):
                    ws[f"{letter}{r}"].number_format = "yyyy-mm-dd"

        # Drop-down lists for the validated columns
        def add_dropdown(col_letter, options):
            dv = DataValidation(type="list",
                                formula1='"' + ",".join(options) + '"',
                                allow_blank=True, showDropDown=False)
            dv.error = "Please pick a value from the list."
            dv.errorTitle = "Invalid value"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        add_dropdown(get_column_letter(col_index["status"]),
                     ["todo", "just", "progress", "hold", "blocked", "review",
                      "pushed_dev", "tested_dev", "pushed_uat", "tested_uat",
                      "pushed_cloud", "tested_cloud", "done"])
        add_dropdown(get_column_letter(col_index["priority"]),
                     ["high", "medium", "low"])
        add_dropdown(get_column_letter(col_index["quadrant"]),
                     ["q1", "q2", "q3", "q4"])
        add_dropdown(get_column_letter(col_index["type"]),
                     ["task", "issue"])

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="taskflow_template.xlsx",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: EDIT TASK ───────────────────────────────────────────────
@app.route("/api/tasks/<task_id>", methods=["PUT"])
def edit_task(task_id):
    try:
        data = request.get_json()
        with cache_lock:
            current = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not current:
            return jsonify({"error": "Task not found."}), 404

        updated = dict(current)
        updated["updated_at"] = get_timestamp()
        updatable = ["title","description","assigned_to","team","due_date","priority","status","sub_category","quadrant","notes","type","start_date","attachments"]
        for f in updatable:
            if f in data:
                val = data[f]
                if f in ("due_date", "start_date"):
                    updated[f] = normalize_date(val)
                else:
                    updated[f] = val.strip() if isinstance(val, str) else val

        if updated["priority"] not in VALID_PRIORITIES: updated["priority"] = "medium"
        if updated["status"] not in VALID_STATUSES: updated["status"] = "todo"
        if updated["quadrant"] not in VALID_QUADRANTS: updated["quadrant"] = "q2"

        with cache_lock:
            cache["tasks"] = [updated if t["task_id"] == task_id else t for t in cache["tasks"]]

        def db_write():
            supabase.table("tasks").update(updated).eq("task_id", task_id).execute()
        write_in_background(db_write)
        changed = [f for f in updatable if f in data]
        log_activity(data.get("user", "System"), "Edited", task_id, updated["title"], f"Changed: {', '.join(changed)}")
        return jsonify({"message": "Task updated successfully.", "task": updated}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: TOGGLE STATUS ───────────────────────────────────────────
@app.route("/api/tasks/<task_id>/toggle", methods=["PATCH"])
def toggle_task(task_id):
    try:
        data = request.get_json(silent=True) or {}
        req_user = data.get("user", "System")
        with cache_lock:
            current = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not current:
            return jsonify({"error": "Task not found."}), 404
        new_status = "done" if current["status"] != "done" else "todo"
        now = get_timestamp()
        with cache_lock:
            for t in cache["tasks"]:
                if t["task_id"] == task_id:
                    t["status"] = new_status
                    t["updated_at"] = now
        def db_write():
            supabase.table("tasks").update({"status": new_status, "updated_at": now}).eq("task_id", task_id).execute()
        write_in_background(db_write)
        log_activity(req_user, "Status Changed", task_id, current["title"], f"Status: {new_status}")
        return jsonify({"message": f"Task marked as {new_status}.", "task_id": task_id, "status": new_status}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: DUPLICATE TASK ──────────────────────────────────────────
@app.route("/api/tasks/<task_id>/duplicate", methods=["POST"])
def duplicate_task(task_id):
    try:
        data = request.get_json(silent=True) or {}
        req_user = data.get("user", "System")
        with cache_lock:
            original = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not original:
            return jsonify({"error": "Task not found."}), 404
        now = get_timestamp()
        new_task = dict(original)
        new_task["task_id"] = generate_task_id()
        new_task["title"] = f"{original['title']} (copy)"
        new_task["status"] = "todo"
        new_task["created_at"] = now
        new_task["updated_at"] = now
        with cache_lock:
            cache["tasks"].insert(0, new_task)
        def db_write():
            supabase.table("tasks").insert(new_task).execute()
        write_in_background(db_write)
        log_activity(req_user, "Duplicated", new_task["task_id"], new_task["title"], f"From: {task_id}")
        return jsonify({"message": "Task duplicated.", "task": new_task}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: DELETE TASK ─────────────────────────────────────────────
@app.route("/api/tasks/<task_id>", methods=["DELETE"])
def delete_task(task_id):
    try:
        req_user = request.args.get("user", "System")
        with cache_lock:
            deleted = next((t for t in cache["tasks"] if t["task_id"] == task_id), None)
        if not deleted:
            return jsonify({"error": "Task not found."}), 404
        with cache_lock:
            cache["tasks"] = [t for t in cache["tasks"] if t["task_id"] != task_id]
        def db_write():
            supabase.table("tasks").delete().eq("task_id", task_id).execute()
        write_in_background(db_write)
        log_activity(req_user, "Deleted", task_id, deleted.get("title", ""), "")
        return jsonify({"message": "Task deleted.", "task_id": task_id}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: BULK UPDATE ─────────────────────────────────────────────
@app.route("/api/tasks/bulk-update", methods=["PATCH"])
def bulk_update_tasks():
    try:
        data = request.get_json()
        task_ids = data.get("task_ids", [])
        field = data.get("field", "").strip()
        value = data.get("value", "").strip()
        if not task_ids or not field:
            return jsonify({"error": "task_ids and field are required."}), 400
        allowed = ["status","priority","quadrant","assigned_to","team","sub_category"]
        if field not in allowed:
            return jsonify({"error": f"Field must be one of: {', '.join(allowed)}"}), 400
        now = get_timestamp()
        with cache_lock:
            for t in cache["tasks"]:
                if t["task_id"] in task_ids:
                    t[field] = value
                    t["updated_at"] = now
        def db_write():
            supabase.table("tasks").update({field: value, "updated_at": now}).in_("task_id", task_ids).execute()
        write_in_background(db_write)
        log_activity(data.get("user", "System"), "Bulk Updated", "", f"{len(task_ids)} tasks", f"Set {field} = {value}")
        return jsonify({"message": f"Updated {len(task_ids)} tasks.", "count": len(task_ids)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: BULK DELETE ─────────────────────────────────────────────
@app.route("/api/tasks/bulk-delete", methods=["POST"])
def bulk_delete_tasks():
    try:
        data = request.get_json()
        task_ids = data.get("task_ids", [])
        if not task_ids:
            return jsonify({"error": "task_ids is required."}), 400
        with cache_lock:
            cache["tasks"] = [t for t in cache["tasks"] if t["task_id"] not in task_ids]
        def db_write():
            supabase.table("tasks").delete().in_("task_id", task_ids).execute()
        write_in_background(db_write)
        log_activity(data.get("user", "System"), "Bulk Deleted", "", f"{len(task_ids)} tasks", "")
        return jsonify({"message": f"Deleted {len(task_ids)} tasks.", "count": len(task_ids)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: ACTIVITY LOG ────────────────────────────────────────────
@app.route("/api/activity", methods=["GET"])
def get_activity():
    try:
        limit = int(request.args.get("limit", 50))
        with cache_lock:
            activity = list(cache.get("activity", []))[:limit]
        return jsonify({"activity": activity}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: AUTO-ASSIGN ─────────────────────────────────────────────
@app.route("/api/auto-assign", methods=["POST"])
def auto_assign():
    try:
        data = request.get_json()
        titles = data.get("titles", [])
        if not titles:
            return jsonify({"error": "No titles provided."}), 400
        if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
            return jsonify({"error": "Groq API key not configured."}), 500

        with cache_lock:
            all_tasks = list(cache["tasks"])
        examples = [{"title": t["title"], "function": t["team"], "sub_category": t["sub_category"]} for t in all_tasks if t["team"] and t["sub_category"]][:50]

        prompt = f"""You are a task classifier for LingoTran, a language learning platform.
AVAILABLE FUNCTIONS AND SUB-CATEGORIES:
{json.dumps(FUNCTIONS, indent=2)}
EXAMPLES FROM EXISTING TASKS:
{json.dumps(examples, indent=2)}
Classify each new task title into the best matching function and sub_category.
Respond ONLY with a JSON array. Each item must have: title, function, sub_category
New task titles: {json.dumps(titles)}"""

        response = http_requests.post("https://api.groq.com/openai/v1/chat/completions",
            json={"model":"llama-3.1-8b-instant","messages":[{"role":"system","content":"JSON-only classifier."},{"role":"user","content":prompt}],"temperature":0.2,"max_tokens":2000},
            headers={"Authorization":f"Bearer {GROQ_API_KEY}","Content-Type":"application/json"}, timeout=30)
        answer = response.json()["choices"][0]["message"]["content"].strip()
        if answer.startswith("```"):
            answer = answer.split("\n",1)[-1].rsplit("```",1)[0].strip()
        assignments = json.loads(answer)
        for a in assignments:
            if a.get("function") not in FUNCTIONS: a["function"] = ""
            if a.get("sub_category") not in FUNCTIONS.get(a.get("function",[]),[]): 
                subs = FUNCTIONS.get(a.get("function"),[])
                a["sub_category"] = subs[0] if subs else ""
        return jsonify({"assignments": assignments}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: AI INSIGHTS ─────────────────────────────────────────────
@app.route("/api/insights", methods=["GET"])
def get_insights():
    try:
        if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
            return jsonify({"insights": []}), 200
        with cache_lock:
            all_tasks = list(cache["tasks"])
        today = datetime.now().strftime("%Y-%m-%d")
        open_tasks = [t for t in all_tasks if t["status"] != "done"]
        if not open_tasks:
            return jsonify({"insights": []}), 200
        task_lines = [f"Title:{t['title']} | Owner:{t['assigned_to'] or 'Unassigned'} | Function:{t['team']} | Priority:{t['priority']} | Status:{t['status']} | Due:{t['due_date'] or 'No date'} {'OVERDUE' if t['due_date'] and t['due_date'] < today else ''}" for t in open_tasks]
        total=len(all_tasks); done=sum(1 for t in all_tasks if t["status"]=="done"); overdue=sum(1 for t in all_tasks if t["status"]!="done" and t["due_date"] and t["due_date"]<today); blocked=sum(1 for t in all_tasks if t["status"] in ["blocked","hold"])
        workload = {}
        for t in all_tasks:
            if t["status"]!="done" and t["assigned_to"] and t["assigned_to"]!="Unassigned":
                workload[t["assigned_to"]]=workload.get(t["assigned_to"],0)+1
        prompt = f"""You are a smart task management assistant for LingoTran.
Today: {today}. Stats: Total:{total}|Done:{done}|Overdue:{overdue}|Blocked:{blocked}. Workload: {json.dumps(workload)}
OPEN TASKS:\n{chr(10).join(task_lines[:50])}
Generate exactly 3 short actionable notifications (1 sentence each, max 15 words, different types, add emoji).
Respond ONLY with a JSON array of 3 strings."""
        response = http_requests.post("https://api.groq.com/openai/v1/chat/completions",
            json={"model":"llama-3.1-8b-instant","messages":[{"role":"system","content":"Return only valid JSON arrays of strings."},{"role":"user","content":prompt}],"temperature":0.4,"max_tokens":300},
            headers={"Authorization":f"Bearer {GROQ_API_KEY}","Content-Type":"application/json"}, timeout=15)
        answer = response.json()["choices"][0]["message"]["content"].strip()
        if answer.startswith("```"): answer = answer.split("\n",1)[-1].rsplit("```",1)[0].strip()
        insights = [str(i) for i in json.loads(answer)[:3]]
        return jsonify({"insights": insights}), 200
    except Exception as e:
        print(f"Insights error: {e}")
        return jsonify({"insights": []}), 200

# ── API: ASK AI ──────────────────────────────────────────────────
@app.route("/api/ask", methods=["POST"])
def ask_ai():
    try:
        data = request.get_json()
        question = data.get("question", "").strip()
        if not question:
            return jsonify({"error": "Question is required."}), 400
        if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
            return jsonify({"error": "No AI API key configured."}), 500

        with cache_lock:
            all_tasks = list(cache["tasks"])
            all_users = list(cache["users"])
            all_teams = list(cache["teams"])

        today = datetime.now().strftime("%Y-%m-%d")
        task_lines = [f"ID:{t['task_id']}|Title:{t['title']}|Owner:{t['assigned_to']}|Function:{t['team']}|Priority:{t['priority']}|Status:{t['status']}|Due:{t['due_date'] or 'No date'}{'|OVERDUE' if t['due_date'] and t['status']!='done' and t['due_date']<today else ''}" for t in all_tasks]
        total=len(all_tasks); done=sum(1 for t in all_tasks if t["status"]=="done"); in_progress=sum(1 for t in all_tasks if t["status"] in ["progress","just"]); blocked=sum(1 for t in all_tasks if t["status"] in ["blocked","hold"]); overdue_count=sum(1 for t in all_tasks if t["status"]!="done" and t["due_date"] and t["due_date"]<today)

        system_prompt = f"""You are TaskFlow AI Assistant. Today: {today}.
TEAM MEMBERS: {', '.join(u['name']+'('+u['team']+')' for u in all_users)}
STATS: Total:{total}|Done:{done}|In Progress:{in_progress}|Blocked:{blocked}|Overdue:{overdue_count}
ALL TASKS:\n{chr(10).join(task_lines)}
INSTRUCTIONS: Only answer questions about tasks, team workload, deadlines, and project management. 
If asked anything unrelated say "I can only help with task-related questions."
Be concise and data-driven."""

        response = http_requests.post("https://api.groq.com/openai/v1/chat/completions",
            json={"model":"llama-3.1-8b-instant","messages":[{"role":"system","content":system_prompt},{"role":"user","content":question}],"temperature":0.3,"max_tokens":1000},
            headers={"Authorization":f"Bearer {GROQ_API_KEY}","Content-Type":"application/json"}, timeout=30)
        answer = response.json()["choices"][0].get("message",{}).get("content","")
        return jsonify({"answer": answer or "Sorry, I could not generate a response."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: REFRESH ─────────────────────────────────────────────────
@app.route("/api/refresh", methods=["POST"])
def refresh_cache():
    sync_cache_from_supabase()
    return jsonify({"message": "Cache refreshed.", "last_sync": cache["last_sync"]}), 200

# ── STARTUP ──────────────────────────────────────────────────────
def startup():
    print("Seeding default teams...")
    seed_teams()
    print("Loading data into cache...")
    sync_cache_from_supabase()
    sync_thread = threading.Thread(target=background_sync, daemon=True)
    sync_thread.start()
    print("Background sync started.")

startup()

if __name__ == "__main__":
    print("Starting TaskFlow on http://127.0.0.1:5000")
    app.run(debug=True, host="0.0.0.0", port=5000, use_reloader=False)